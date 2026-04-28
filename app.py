
import streamlit as st
import streamlit.components.v1 as components
import pandas as pd
import numpy as np
from pathlib import Path
import pickle
import json
import html
import hashlib
import hmac
import secrets
import time
import base64
import re
import os
import io
import zipfile
import unicodedata
from datetime import datetime
from urllib.parse import quote, urlencode
from openpyxl import load_workbook

# ====================== CONFIG ======================

st.set_page_config(
    page_title="Espace Commissions ECOHABITAT",
    layout="wide",
    initial_sidebar_state="collapsed"
)

# ====================== CSS DESIGN ======================

st.markdown("""
<style>

/* GLOBAL */
.stApp {
    background: #F6F8F4;
    color: #1F2933;
}

html, body, [class*="css"] {
    color: #1F2933 !important;
}

/* SIDEBAR */
[data-testid="stSidebar"] {
    background: #FFFFFF;
    border-right: 1px solid #E5E5E5;
}

[data-testid="stSidebar"] * {
    color: #1F2933 !important;
}

/* TEXTES */
h1, h2, h3, h4, h5, h6,
p, label {
    color: #1F2933 !important;
}

/* MARKDOWN + TABLEAUX */
[data-testid="stMarkdownContainer"] * {
    color: #1F2933 !important;
}

[data-testid="stDataFrame"] * {
    color: inherit !important;
}

/* HEADER */
.eco-header {
    background: white;
    padding: 22px;
    border-radius: 16px;
    box-shadow: 0 3px 12px rgba(0,0,0,0.06);
    margin-bottom: 20px;
    border-left: 6px solid #66B32E;
}

.eco-title {
    font-size: 32px;
    font-weight: 800;
    color: #1F2933 !important;
}

.eco-subtitle {
    color: #66B32E !important;
    font-size: 15px;
    font-weight: 600;
}

/* TABS */
button[data-baseweb="tab"] {
    font-weight: 700;
    color: #1F2933 !important;
}

button[data-baseweb="tab"][aria-selected="true"] {
    color: #66B32E !important;
}

/* BOUTONS */
.stButton > button {
    border-radius: 10px;
    font-weight: 600;
    background: #1F2933;
    color: white !important;
}

.stButton > button[kind="primary"] {
    background-color: #66B32E !important;
}

/* ALERTES */
[data-testid="stAlert"] * {
    color: #1F2933 !important;
}

/* TABLEAUX */
[data-testid="stDataFrame"] {
    border-radius: 12px;
    overflow: hidden;
}

/* CARDS */
.eco-card {
    background: white;
    padding: 20px;
    border-radius: 14px;
    box-shadow: 0 3px 10px rgba(0,0,0,0.06);
    border-left: 6px solid #66B32E;
    text-align: center;
    min-height: 105px;
}

.eco-card-title {
    font-size: 14px;
    color: #6B7280 !important;
    font-weight: 600;
}

.eco-card-value {
    font-size: clamp(18px, 1.45vw, 24px);
    font-weight: 800;
    color: #1F2933 !important;
    margin-top: 8px;
    white-space: nowrap;
}

.commission-summary {
    display: flex;
    align-items: stretch;
    justify-content: center;
    gap: 10px;
    flex-wrap: wrap;
    margin: 8px auto 18px auto;
    width: 100%;
}

.commission-pill {
    background: #FFFFFF;
    border: 1px solid #E5E7EB;
    border-radius: 10px;
    padding: 10px 14px;
    box-shadow: 0 2px 8px rgba(0,0,0,0.04);
}

.commission-pill {
    min-width: 190px;
    text-align: center;
}

.commission-pill span {
    display: block;
    color: #6B7280 !important;
    font-size: 12px;
    font-weight: 700;
}

.commission-pill strong {
    display: block;
    margin-top: 3px;
    font-size: 18px;
    color: #1F2933 !important;
}

.commission-pill.points {
    border-color: #FCD34D;
}

.commission-pill.final {
    border-color: #66B32E;
    background: #F0F9EB;
}

/* CORRECTION BOUTONS SIDEBAR */
[data-testid="stSidebar"] .stButton > button {
    background: #FFFFFF !important;
    color: #1F2933 !important;
    border: 1px solid #D1D5DB !important;
}

[data-testid="stSidebar"] .stButton > button:hover {
    background: #66B32E !important;
    color: white !important;
    border: 1px solid #66B32E !important;
}

[data-testid="stSidebar"] .stButton > button * {
    color: inherit !important;
}

/* INPUT LOGIN / TEXT INPUT */
input[type="text"],
input[type="password"],
textarea,
select {
    background-color: #FFFFFF !important;
    color: #1F2933 !important;
    border: 1px solid #D1D5DB !important;
    border-radius: 8px !important;
}

input:focus, textarea:focus, select:focus {
    border-color: #66B32E !important;
    box-shadow: 0 0 0 1px #66B32E !important;
}

::placeholder {
    color: #9CA3AF !important;
}

</style>
""", unsafe_allow_html=True)


# ====================== STOCKAGE LOCAL / RENDER / OVH ======================

DATA_DIR = Path("/data") if Path("/data").exists() else Path(".")

HISTORIQUE_DIR = DATA_DIR / "historique"
HISTORIQUE_DIR.mkdir(parents=True, exist_ok=True)

USERS_FILE = DATA_DIR / "users.json"
ATTENTE_MOTIFS_FILE = DATA_DIR / "motifs_attente.json"
SETTINGS_FILE = DATA_DIR / "settings.json"
MAX_LOGIN_ATTEMPTS = 5
LOGIN_LOCK_SECONDS = 5 * 60


# ====================== USERS ======================

def hash_password(password):
    salt = secrets.token_hex(16)
    digest = hashlib.pbkdf2_hmac(
        "sha256",
        str(password).encode("utf-8"),
        salt.encode("utf-8"),
        180000
    ).hex()
    return f"pbkdf2_sha256$180000${salt}${digest}"


def verify_password(password, stored_value):
    stored_value = str(stored_value or "")
    if stored_value.startswith("pbkdf2_sha256$"):
        try:
            _, iterations, salt, expected = stored_value.split("$", 3)
            digest = hashlib.pbkdf2_hmac(
                "sha256",
                str(password).encode("utf-8"),
                salt.encode("utf-8"),
                int(iterations)
            ).hex()
            return hmac.compare_digest(digest, expected)
        except Exception:
            return False

    return hmac.compare_digest(stored_value, str(password))


def password_is_hashed(user):
    return str(user.get("password_hash", "")).startswith("pbkdf2_sha256$")


def password_is_strong_enough(password):
    password = str(password or "")
    return (
        len(password) >= 8
        and any(c.isalpha() for c in password)
        and any(c.isdigit() for c in password)
    )


def generate_totp_secret():
    return base64.b32encode(secrets.token_bytes(20)).decode("utf-8").rstrip("=")


def hotp(secret, counter, digits=6):
    normalized_secret = str(secret or "").replace(" ", "").upper()
    padding = "=" * ((8 - len(normalized_secret) % 8) % 8)
    key = base64.b32decode(normalized_secret + padding)
    counter_bytes = int(counter).to_bytes(8, "big")
    digest = hmac.new(key, counter_bytes, hashlib.sha1).digest()
    offset = digest[-1] & 0x0F
    code = int.from_bytes(digest[offset:offset + 4], "big") & 0x7FFFFFFF
    return str(code % (10 ** digits)).zfill(digits)


def verify_totp(secret, code, period=30, window=1):
    code = re.sub(r"\D", "", str(code or ""))
    if len(code) != 6 or not secret:
        return False

    current_counter = int(time.time() // period)
    for drift in range(-window, window + 1):
        if hmac.compare_digest(hotp(secret, current_counter + drift), code):
            return True
    return False


def totp_uri(username, secret, issuer="gesCom EcoHabitat"):
    label = f"{issuer}:{username}"
    params = urlencode({
        "secret": secret,
        "issuer": issuer,
        "algorithm": "SHA1",
        "digits": 6,
        "period": 30,
    })
    return f"otpauth://totp/{quote(label)}?{params}"


def qr_code_url(uri):
    return "https://api.qrserver.com/v1/create-qr-code/?" + urlencode({
        "size": "220x220",
        "data": uri
    })


def get_login_lock_remaining():
    locked_until = st.session_state.get("login_locked_until", 0)
    return max(0, int(locked_until - time.time()))


def register_failed_login():
    st.session_state.login_attempts = st.session_state.get("login_attempts", 0) + 1
    if st.session_state.login_attempts >= MAX_LOGIN_ATTEMPTS:
        st.session_state.login_locked_until = time.time() + LOGIN_LOCK_SECONDS


def reset_login_security_state():
    st.session_state.login_attempts = 0
    st.session_state.login_locked_until = 0


def load_users():
    if USERS_FILE.exists():
        with open(USERS_FILE, "r", encoding="utf-8") as f:
            return json.load(f)

    default_users = {
        "joseph": {
            "password_hash": hash_password("admin123"),
            "role": "admin",
            "nom": "LUCCHINI JOSEPH",
            "agence": None
        }
    }

    save_users(default_users)
    return default_users


def save_users(users):
    with open(USERS_FILE, "w", encoding="utf-8") as f:
        json.dump(users, f, indent=4, ensure_ascii=False)


def load_settings():
    if SETTINGS_FILE.exists():
        with open(SETTINGS_FILE, "r", encoding="utf-8") as f:
            try:
                data = json.load(f)
                return data if isinstance(data, dict) else {}
            except json.JSONDecodeError:
                return {}
    return {}


def save_settings(settings):
    with open(SETTINGS_FILE, "w", encoding="utf-8") as f:
        json.dump(settings, f, indent=4, ensure_ascii=False)


# ====================== MOTIFS D'ATTENTE ======================

MOTIFS_ATTENTE_OPTIONS = [
    "",
    "Pièce administrative manquante",
    "Financement en attente",
    "Acompte manquant",
    "Validation client en attente",
    "Métré / technique à confirmer",
    "OPC / offre commerciale à vérifier",
    "Erreur ou information ProDevis",
    "Autre"
]


def load_motifs_attente():
    if ATTENTE_MOTIFS_FILE.exists():
        with open(ATTENTE_MOTIFS_FILE, "r", encoding="utf-8") as f:
            try:
                data = json.load(f)
                return data if isinstance(data, dict) else {}
            except json.JSONDecodeError:
                return {}
    return {}


def save_motifs_attente(data):
    with open(ATTENTE_MOTIFS_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=4, ensure_ascii=False)


# ====================== FONCTIONS DESIGN ======================

def card(title, value):
    value = str(value).replace(" €", "&nbsp;€").replace(" %", "&nbsp;%")
    st.markdown(f"""
    <div class="eco-card">
        <div class="eco-card-title">{title}</div>
        <div class="eco-card-value">{value}</div>
    </div>
    """, unsafe_allow_html=True)


def commission_summary_bar(base_pct, points_perdus, commission_pct):
    st.markdown(f"""
    <div class="commission-summary">
        <div class="commission-pill">
            <span>📐 Base commission</span>
            <strong>{base_pct:,.0f} %</strong>
        </div>
        <div class="commission-pill points">
            <span>⚠️ Points perdus</span>
            <strong>{points_perdus:,.0f}</strong>
        </div>
        <div class="commission-pill final">
            <span>✅ Commission définitive</span>
            <strong>{commission_pct:,.0f} %</strong>
        </div>
    </div>
    """, unsafe_allow_html=True)


def render_table_with_status_tooltips(df, tooltip_by_index=None, height=600):
    tooltip_by_index = tooltip_by_index or {}
    money_cols = {"TOTAL VENTE", "Vente HT hors acompte", "Bonus / Malus", "CA global affaire"}
    percent_cols = {"Remise %"}

    def format_cell(col, value):
        if pd.isna(value):
            return ""
        if col in money_cols:
            return f"{to_float(value):,.2f} €"
        if col in percent_cols:
            return f"{to_float(value):.2f} %"
        return str(value)

    def cell_class(col, value):
        if col == "Bonus / Malus":
            v = to_float(value)
            if v > 0:
                return " bm-positive"
            if v < 0:
                return " bm-negative"
        if col in money_cols or col in percent_cols:
            return " numeric"
        return ""

    headers = [""] + list(df.columns)
    html_parts = [
        f"""
        <style>
        .eco-table-wrap {{
            max-height: {height}px;
            overflow: auto;
            border: 1px solid #E5E7EB;
            border-radius: 10px;
            background: #FFFFFF;
        }}
        .eco-table {{
            border-collapse: collapse;
            width: max-content;
            min-width: 100%;
            font-family: "Source Sans Pro", -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
            font-size: 14px;
            line-height: 1.35;
        }}
        .eco-table th, .eco-table td {{
            border-bottom: 1px solid #E5E7EB;
            border-right: 1px solid #E5E7EB;
            padding: 9px 10px;
            white-space: nowrap;
            color: #1F2933;
        }}
        .eco-table th {{
            position: sticky;
            top: 0;
            z-index: 2;
            background: #F8FAFC;
            color: #6B7280;
            font-weight: 500;
            text-align: left;
        }}
        .eco-table .row-index {{
            color: #6B7280;
            text-align: right;
            min-width: 34px;
        }}
        .eco-table .numeric {{
            text-align: right;
        }}
        .eco-table .bm-positive {{
            background: #D1FADF;
            color: #065F46;
            font-weight: 700;
            text-align: right;
        }}
        .eco-table .bm-negative {{
            background: #FECACA;
            color: #991B1B;
            font-weight: 700;
            text-align: right;
        }}
        .status-tip {{
            position: relative;
            display: inline-block;
            cursor: help;
        }}
        .status-tip::after {{
            content: attr(data-tip);
            position: absolute;
            left: 0;
            bottom: 145%;
            display: none;
            min-width: 220px;
            max-width: 360px;
            padding: 9px 11px;
            border-radius: 8px;
            background: #1F2933;
            color: #FFFFFF;
            white-space: normal;
            box-shadow: 0 8px 22px rgba(0,0,0,0.18);
            z-index: 20;
        }}
        .status-tip:hover::after {{
            display: block;
        }}
        </style>
        <div class="eco-table-wrap">
        <table class="eco-table">
        """
    ]
    html_parts.append("<thead><tr>")
    for header in headers:
        html_parts.append(f"<th>{html.escape(str(header))}</th>")
    html_parts.append("</tr></thead><tbody>")

    for display_idx, (idx, row) in enumerate(df.iterrows()):
        html_parts.append("<tr>")
        html_parts.append(f'<td class="row-index">{display_idx}</td>')
        for col in df.columns:
            value = row.get(col, "")
            text = format_cell(col, value)
            cls = cell_class(col, value)
            if col == "Statut" and clean_visible(tooltip_by_index.get(idx, "")):
                tip = html.escape(clean_visible(tooltip_by_index.get(idx, "")), quote=True)
                cell_html = f'<span class="status-tip" data-tip="{tip}">⏳ En attente ⓘ</span>'
            else:
                cell_html = html.escape(text)
            html_parts.append(f'<td class="{cls.strip()}">{cell_html}</td>')
        html_parts.append("</tr>")

    html_parts.append("</tbody></table></div>")
    components.html("".join(html_parts), height=height + 40, scrolling=True)


def build_attente_internal_mail_body(df, periode, col_client, col_doc, col_date, col_agence, col_ca_magasin):
    lines = [
        "Bonjour,",
        "",
        f"Point hebdomadaire des dossiers en attente - {periode}",
        "",
        f"Nombre de dossiers : {len(df)}",
    ]

    if col_ca_magasin and col_ca_magasin in df.columns:
        total = pd.to_numeric(df[col_ca_magasin], errors="coerce").fillna(0).sum()
        lines.append(f"CA global en attente : {total:,.2f} EUR")

    lines.append("")
    lines.append(
        "Merci à l'ADV de compléter les motifs manquants et de mettre à jour les motifs déjà renseignés "
        "si la situation du dossier a évolué."
    )
    lines.append("Lien Manager pour compléter les motifs : https://manager.ecohabitat76.fr/")
    lines.append(
        "L'objectif est que chaque dossier en attente indique clairement l'élément bloquant et l'action attendue."
    )
    lines.append("")
    lines.append("Dossiers à suivre :")

    if df.empty:
        lines.append("- Aucun dossier en attente sur la sélection.")
        return "\n".join(lines)

    group_col = col_agence if col_agence and col_agence in df.columns else None
    if group_col:
        grouped_items = df.groupby(df[group_col].fillna("").map(clean_visible), dropna=False)
    else:
        grouped_items = [("Non classé", df)]

    for agence, group in grouped_items:
        lines.append("")
        lines.append(f"Agence : {agence or 'Non renseignée'}")
        for _, row in group.iterrows():
            client = clean_visible(row.get(col_client, "")) if col_client else "Dossier"
            doc = clean_visible(row.get(col_doc, "")) if col_doc else ""
            date_doc = ""
            if col_date and col_date in row.index:
                parsed_date = pd.to_datetime(row.get(col_date), errors="coerce")
                if not pd.isna(parsed_date):
                    date_doc = parsed_date.strftime("%d/%m/%Y")
            commerciaux = clean_visible(row.get("Commerciaux", ""))
            motif = clean_visible(row.get("Motif d'attente", "")) or "Motif à compléter"
            detail = clean_visible(row.get("Détail motif", "")) or "Élément manquant à préciser"

            line = f"- {client}"
            if doc:
                line += f" | Doc {doc}"
            if date_doc:
                line += f" | {date_doc}"
            if commerciaux:
                line += f" | {commerciaux}"
            line += f" | Motif : {motif} | À obtenir : {detail}"
            lines.append(line)

    lines.extend([
        "",
        "Merci également de mettre à jour les dossiers dès réception des éléments afin de permettre leur passage en statut actif.",
        "",
        "Cordialement,"
    ])
    return "\n".join(lines)


# ====================== FONCTIONS OUTILS ======================

def normalize_key(s):
    if pd.isna(s):
        return ""
    t = str(s).replace(chr(160), " ").replace("\t", " ").replace("\n", " ").replace("\r", " ")
    return " ".join(t.strip().upper().split())


def strip_accents(s):
    return "".join(
        c for c in unicodedata.normalize("NFD", str(s))
        if unicodedata.category(c) != "Mn"
    )


def clean_visible(s):
    if pd.isna(s):
        return ""
    return " ".join(str(s).replace(chr(160), " ").split()).strip()


def to_float(value, default=0.0):
    try:
        value = pd.to_numeric(value, errors="coerce")
    except Exception:
        return default
    return default if pd.isna(value) else float(value)


def is_excluded_from_evp(nom):
    k = strip_accents(normalize_key(nom))
    return k in [
        "LAVISSE GUILLAUME",
        "LAVISSE FABIEN",
        "LUCCHINI JOSEPH",
        "LUCCHIN JOSEPH",
        "PETIT LILIAN"
    ]


def is_responsable_agence(nom):
    return strip_accents(normalize_key(nom)) in [
        "AYACHE ADEL",
        "EL GHAZOUANI NAHIM",
        "VUE JONATHAN",
    ]


def resolve_nom_evp(nom):
    k = strip_accents(normalize_key(nom))
    aliases = {
        "LEMMONIER HENRI": "LEMONNIER HENRI",
        "LEMONIER HENRI": "LEMONNIER HENRI",
        "LEMONNIER HENRI": "LEMONNIER HENRI",
        "ELGHAZOUANI NAHIM": "EL GHAZOUANI NAHIM",
        "EL GHAZOUANI NAHIM": "EL GHAZOUANI NAHIM",
        "ADEL AYACHE": "AYACHE ADEL",
        "AYACHE ADEL": "AYACHE ADEL",
        "BALDACCHINO ANTOINE": "BALDACCHINO ANTOINE",
        "AGASSE ESTEBAN": "AGASSE ESTEBAN",
        "DUSSART ROBIN": "DUSSART ROBIN",
        "RENOUT KEVIN": "RENOUT KEVIN",
        "VUE JONATHAN": "VUE JONATHAN",
    }
    return aliases.get(k, k)


def prime_magasin(ca_ht):
    if ca_ht >= 300000:
        return 1000
    elif ca_ht >= 250000:
        return 750
    elif ca_ht >= 200000:
        return 500
    return 0


def calculate_commission(ca_ok, remise_pct):
    if ca_ok < 10000:
        base = 0
    elif ca_ok < 15000:
        base = 3
    elif ca_ok < 20000:
        base = 6
    elif ca_ok < 30000:
        base = 8
    elif ca_ok < 40000:
        base = 9
    elif ca_ok < 50000:
        base = 10
    elif ca_ok < 60000:
        base = 12
    else:
        base = 14

    if remise_pct <= 16:
        points = 0
    elif remise_pct > 25:
        base = 0
        points = 0
    else:
        points = int(np.floor(remise_pct - 15))

    commission_pct = max(0, base - points)
    commission_euro = round(ca_ok * commission_pct / 100, 2)

    return base, points, commission_pct, commission_euro


def row_remise_pct(row, col_rem, col_catalogue):
    if not col_rem or col_rem not in row.index or not col_catalogue or col_catalogue not in row.index:
        return 0.0

    remise = to_float(row.get(col_rem))
    catalogue = to_float(row.get(col_catalogue))
    return (remise / catalogue * 100) if catalogue > 0 else 0.0


def calculate_directeur_commission(detail, col_vente, col_rem, col_catalogue):
    if detail.empty or not col_vente or col_vente not in detail.columns:
        return 0.0, 10, 0, 10, 0.0

    ca_commissionnable = sum_numeric_col(detail, col_vente)
    commission_euro = round(ca_commissionnable * 0.10, 2)
    return round(ca_commissionnable, 2), 10, 0, 10, commission_euro


def count_remise_over_threshold(detail, col_rem, col_catalogue, threshold=30):
    if detail.empty:
        return 0
    return int(
        detail.apply(
            lambda row: row_remise_pct(row, col_rem, col_catalogue) > threshold,
            axis=1
        ).sum()
    )


def find_col(df, includes=None, excludes=None):
    includes = includes or []
    excludes = excludes or []

    for col in df.columns:
        c = str(col).upper()
        if all(i.upper() in c for i in includes) and not any(e.upper() in c for e in excludes):
            return col
    return None


def get_col_by_excel_position(df, index_1_based):
    idx = index_1_based - 1
    if 0 <= idx < len(df.columns):
        return df.columns[idx]
    return None


def is_opc(row, col_op):
    if not col_op or col_op not in row.index:
        return False
    valeur = normalize_key(row.get(col_op, ""))
    return valeur in ["OUI", "YES", "1", "TRUE", "VRAI"]


def vendeur_mask(df, vendeur, colonnes_commerciaux):
    mask = pd.Series(False, index=df.index)
    for col in colonnes_commerciaux:
        if col and col in df.columns:
            mask |= df[col].apply(normalize_key) == normalize_key(vendeur)
    return mask


def list_commerciaux_row(row, colonnes_commerciaux):
    noms = []
    seen = set()

    for col in colonnes_commerciaux:
        if not col or col not in row.index:
            continue

        nom = clean_visible(row.get(col))
        k = normalize_key(nom)

        if nom and k not in seen:
            noms.append(nom)
            seen.add(k)

    return " / ".join(noms)


def count_vendeurs_row(row, colonnes_commerciaux):
    nb = 0
    for col in colonnes_commerciaux:
        if col and col in row.index and clean_visible(row.get(col)):
            nb += 1
    return max(nb, 1)


def calculate_bonus_malus_row(row, col_vente, col_catalogue, col_op, colonnes_commerciaux):
    if not col_vente or col_vente not in row.index or not col_catalogue or col_catalogue not in row.index:
        return 0.0

    vente = pd.to_numeric(row.get(col_vente), errors="coerce")
    vente = float(vente) if pd.notna(vente) else 0.0

    catalogue = pd.to_numeric(row.get(col_catalogue), errors="coerce")
    catalogue = float(catalogue) if pd.notna(catalogue) else 0.0

    objectif_15_par_vendeur = (catalogue * 0.85) / count_vendeurs_row(row, colonnes_commerciaux)
    bonus_malus = vente - objectif_15_par_vendeur

    if is_opc(row, col_op) and bonus_malus < 0:
        return 0.0

    return bonus_malus


def calculate_bonus_malus_by_vendor(df_ok, df_c, col_vente, col_catalogue, col_op, colonnes_commerciaux, key_cols):
    bonus_malus = {}

    def add_to_vendor(row, target_cols):
        montant = calculate_bonus_malus_row(row, col_vente, col_catalogue, col_op, colonnes_commerciaux)

        for col in colonnes_commerciaux:
            if not col or col not in row.index:
                continue

            nom = clean_visible(row.get(col))
            if not nom:
                continue

            k = normalize_key(nom)
            if k not in bonus_malus:
                bonus_malus[k] = {
                    "bonus_malus_ok": 0.0,
                    "bonus_malus_global": 0.0
                }

            for target_col in target_cols:
                bonus_malus[k][target_col] += montant

    if not df_c.empty:
        for _, row in df_c.iterrows():
            add_to_vendor(row, ["bonus_malus_global"])

    if not df_ok.empty:
        for _, row in df_ok.iterrows():
            add_to_vendor(row, ["bonus_malus_ok"])

    return bonus_malus


def calculate_remise_by_vendor(df_ok, df_c, col_catalogue, col_rem, col_op, colonnes_commerciaux):
    remises = {}

    def add_to_vendor(row, prefix):
        catalogue = to_float(row.get(col_catalogue)) if col_catalogue and col_catalogue in row.index else 0.0
        remise = to_float(row.get(col_rem)) if col_rem and col_rem in row.index else 0.0
        remise_sans_op = 0.0 if is_opc(row, col_op) else remise
        catalogue_hors_op = 0.0 if is_opc(row, col_op) else catalogue

        for col in colonnes_commerciaux:
            if not col or col not in row.index:
                continue

            nom = clean_visible(row.get(col))
            if not nom:
                continue

            k = normalize_key(nom)
            if k not in remises:
                remises[k] = {
                    "ok_rem_total": 0.0,
                    "ok_rem_sans_op": 0.0,
                    "ok_catalogue_total": 0.0,
                    "ok_rem_hors_op": 0.0,
                    "ok_catalogue_hors_op": 0.0,
                    "global_rem_total": 0.0,
                    "global_rem_sans_op": 0.0,
                    "global_catalogue_total": 0.0,
                    "global_rem_hors_op": 0.0,
                    "global_catalogue_hors_op": 0.0,
                }

            remises[k][f"{prefix}_rem_total"] += remise
            remises[k][f"{prefix}_rem_sans_op"] += remise_sans_op
            remises[k][f"{prefix}_catalogue_total"] += catalogue
            remises[k][f"{prefix}_rem_hors_op"] += remise_sans_op
            remises[k][f"{prefix}_catalogue_hors_op"] += catalogue_hors_op

    if not df_c.empty:
        for _, row in df_c.iterrows():
            add_to_vendor(row, "global")

    if not df_ok.empty:
        for _, row in df_ok.iterrows():
            add_to_vendor(row, "ok")

    return remises


def calculate_remise_pct(df, col_catalogue, col_rem, col_op):
    if df.empty or not col_catalogue or col_catalogue not in df.columns or not col_rem or col_rem not in df.columns:
        return 0.0

    catalogue = pd.to_numeric(df[col_catalogue], errors="coerce").fillna(0)
    remise = pd.to_numeric(df[col_rem], errors="coerce").fillna(0)

    if col_op and col_op in df.columns:
        opc_mask = df.apply(lambda row: is_opc(row, col_op), axis=1)
        remise = remise.mask(opc_mask, 0)
        catalogue = catalogue.mask(opc_mask, 0)

    total_catalogue = catalogue.sum()
    return round(remise.sum() / total_catalogue * 100, 2) if total_catalogue > 0 else 0.0


def count_opc_rows(df, col_op):
    if df.empty or not col_op or col_op not in df.columns:
        return 0
    return int(df.apply(lambda row: is_opc(row, col_op), axis=1).sum())


def ensure_remise_columns(df):
    if df.empty:
        return df

    df = df.copy()
    if "remise_ok_pct" not in df.columns:
        df["remise_ok_pct"] = df.get("remise_hors_opc_pct", 0.0)
    if "remise_global_pct" not in df.columns:
        df["remise_global_pct"] = df["remise_ok_pct"]
    if "remise_hors_opc_global_pct" not in df.columns:
        df["remise_hors_opc_global_pct"] = df.get("remise_hors_opc_pct", 0.0)

    for col in ["remise_ok_pct", "remise_global_pct", "remise_hors_opc_pct", "remise_hors_opc_global_pct"]:
        if col not in df.columns:
            df[col] = 0.0
        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0.0)

    return df


def recompute_df_vendeurs_indicators(
    df_vendeurs,
    df_ok,
    df_c,
    col_vente,
    col_catalogue,
    col_rem,
    col_op,
    colonnes_commerciaux,
    key_cols,
    col_client=None,
    col_agence=None,
    col_ca_magasin=None
):
    if df_vendeurs.empty:
        return df_vendeurs

    df_vendeurs = ensure_bonus_malus_columns(df_vendeurs)
    df_vendeurs = ensure_remise_columns(df_vendeurs)
    bonus_malus = calculate_bonus_malus_by_vendor(
        df_ok,
        df_c,
        col_vente,
        col_catalogue,
        col_op,
        colonnes_commerciaux,
        key_cols
    )
    remises = calculate_remise_by_vendor(df_ok, df_c, col_catalogue, col_rem, col_op, colonnes_commerciaux)

    for idx, row in df_vendeurs.iterrows():
        k = normalize_key(row.get("Commercial"))
        values = bonus_malus.get(k, {})
        df_vendeurs.at[idx, "bonus_malus_ok"] = round(values.get("bonus_malus_ok", 0.0), 2)
        df_vendeurs.at[idx, "bonus_malus_global"] = round(values.get("bonus_malus_global", 0.0), 2)

        ok_detail = df_ok[vendeur_mask(df_ok, row.get("Commercial"), colonnes_commerciaux)].copy()
        attente_detail = df_c[vendeur_mask(df_c, row.get("Commercial"), colonnes_commerciaux)].copy()
        attente_detail = remove_attente_already_ok(
            ok_detail,
            attente_detail,
            key_cols,
            col_client,
            col_agence,
            col_vente,
            col_ca_magasin,
            col_catalogue
        )
        ca_ok = sum_numeric_col(ok_detail, col_vente)
        ca_attente = sum_numeric_col(attente_detail, col_vente)
        nb_ok = len(ok_detail)
        nb_total = len(ok_detail) + len(attente_detail)
        detail_global = pd.concat([ok_detail, attente_detail], ignore_index=True)
        nb_opc_total = count_opc_rows(detail_global, col_op)
        ratio_opc_total = round(nb_opc_total / nb_total * 100, 2) if nb_total > 0 else 0.0

        df_vendeurs.at[idx, "ca_ok"] = round(ca_ok, 2)
        df_vendeurs.at[idx, "ca_attente"] = round(ca_attente, 2)
        df_vendeurs.at[idx, "ca_total"] = round(ca_ok + ca_attente, 2)
        df_vendeurs.at[idx, "nb_ok"] = nb_ok
        df_vendeurs.at[idx, "nb_total"] = nb_total
        df_vendeurs.at[idx, "nb_opc_total"] = nb_opc_total
        df_vendeurs.at[idx, "ratio_opc_total_pct"] = ratio_opc_total

        if is_responsable_agence(row.get("Commercial")):
            ca_commissionnable_ok, base_comm, points, comm_def, euro = calculate_directeur_commission(
                ok_detail,
                col_vente,
                col_rem,
                col_catalogue
            )
            ca_commissionnable_total, _, _, _, _ = calculate_directeur_commission(
                detail_global,
                col_vente,
                col_rem,
                col_catalogue
            )
            df_vendeurs.at[idx, "ca_commissionnable_directeur_ok"] = ca_commissionnable_ok
            df_vendeurs.at[idx, "ca_commissionnable_directeur_total"] = ca_commissionnable_total
            df_vendeurs.at[idx, "nb_remise_plus_30_ok"] = count_remise_over_threshold(
                ok_detail,
                col_rem,
                col_catalogue
            )
            df_vendeurs.at[idx, "nb_remise_plus_30_total"] = count_remise_over_threshold(
                detail_global,
                col_rem,
                col_catalogue
            )
            df_vendeurs.at[idx, "base_commission_pct"] = base_comm
            df_vendeurs.at[idx, "points_perdus"] = points
            df_vendeurs.at[idx, "commission_pct"] = comm_def
            df_vendeurs.at[idx, "commission_eur"] = euro

        remise_values = remises.get(k, {})
        ok_catalogue_total = remise_values.get("ok_catalogue_total", 0.0)
        global_catalogue_total = remise_values.get("global_catalogue_total", 0.0)
        ok_catalogue_hors_op = remise_values.get("ok_catalogue_hors_op", 0.0)
        global_catalogue_hors_op = remise_values.get("global_catalogue_hors_op", 0.0)

        df_vendeurs.at[idx, "remise_ok_pct"] = round(
            remise_values.get("ok_rem_sans_op", 0.0) / ok_catalogue_hors_op * 100,
            2
        ) if ok_catalogue_hors_op > 0 else 0.0
        df_vendeurs.at[idx, "remise_global_pct"] = round(
            remise_values.get("global_rem_sans_op", 0.0) / global_catalogue_hors_op * 100,
            2
        ) if global_catalogue_hors_op > 0 else 0.0
        df_vendeurs.at[idx, "remise_hors_opc_pct"] = round(
            remise_values.get("ok_rem_hors_op", 0.0) / ok_catalogue_hors_op * 100,
            2
        ) if ok_catalogue_hors_op > 0 else 0.0
        df_vendeurs.at[idx, "remise_hors_opc_global_pct"] = round(
            remise_values.get("global_rem_hors_op", 0.0) / global_catalogue_hors_op * 100,
            2
        ) if global_catalogue_hors_op > 0 else 0.0

    return df_vendeurs


def ensure_bonus_malus_columns(df):
    if df.empty:
        return df

    df = df.copy()
    for col in ["bonus_malus_ok", "bonus_malus_global"]:
        if col not in df.columns:
            df[col] = 0.0
        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0.0)
    return df


def agence_mask(df, agence, col_agence):
    if not col_agence or col_agence not in df.columns:
        return pd.Series(False, index=df.index)
    return df[col_agence].apply(normalize_key) == normalize_key(agence)


def make_affaire_key(row, key_cols):
    return "|".join(normalize_key(row.get(c, "")) for c in key_cols)


def amount_key(row, col):
    if not col or col not in row.index:
        return "0.00"
    return f"{to_float(row.get(col)):,.2f}"


def make_affaire_match_key(row, col_client, col_agence, col_vente, col_ca_magasin, col_catalogue):
    parts = [
        normalize_key(row.get(col_client, "")) if col_client else "",
        normalize_key(row.get(col_agence, "")) if col_agence else "",
        amount_key(row, col_vente),
        amount_key(row, col_ca_magasin),
        amount_key(row, col_catalogue),
    ]
    return "|".join(parts)


def make_affaire_client_agence_key(row, col_client, col_agence):
    return "|".join([
        normalize_key(row.get(col_client, "")) if col_client else "",
        normalize_key(row.get(col_agence, "")) if col_agence else "",
    ])


def make_attente_tracking_key(row, col_client, col_doc, col_agence, col_ca_magasin, col_catalogue):
    parts = [
        normalize_key(row.get(col_doc, "")) if col_doc else "",
        normalize_key(row.get(col_client, "")) if col_client else "",
        normalize_key(row.get(col_agence, "")) if col_agence else "",
        amount_key(row, col_ca_magasin),
        amount_key(row, col_catalogue),
    ]
    return "|".join(parts)


def remove_attente_already_ok(ok_detail, attente_detail, key_cols, col_client, col_agence, col_vente, col_ca_magasin, col_catalogue):
    if ok_detail.empty or attente_detail.empty:
        return attente_detail

    ok_detail = ok_detail.copy()
    attente_detail = attente_detail.copy()

    ok_detail["_AFFAIRE_MATCH_KEY_"] = ok_detail.apply(
        lambda row: make_affaire_match_key(row, col_client, col_agence, col_vente, col_ca_magasin, col_catalogue),
        axis=1
    )
    attente_detail["_AFFAIRE_MATCH_KEY_"] = attente_detail.apply(
        lambda row: make_affaire_match_key(row, col_client, col_agence, col_vente, col_ca_magasin, col_catalogue),
        axis=1
    )

    ok_match_keys = set(ok_detail["_AFFAIRE_MATCH_KEY_"])
    attente_detail = attente_detail[~attente_detail["_AFFAIRE_MATCH_KEY_"].isin(ok_match_keys)].copy()

    ok_detail = ok_detail.drop(columns=["_AFFAIRE_MATCH_KEY_"], errors="ignore")
    attente_detail = attente_detail.drop(columns=["_AFFAIRE_MATCH_KEY_"], errors="ignore")

    if key_cols and not attente_detail.empty:
        ok_detail["_AFFAIRE_KEY_"] = ok_detail.apply(lambda row: make_affaire_key(row, key_cols), axis=1)
        attente_detail["_AFFAIRE_KEY_"] = attente_detail.apply(lambda row: make_affaire_key(row, key_cols), axis=1)

        ok_keys = set(ok_detail["_AFFAIRE_KEY_"])
        attente_detail = attente_detail[~attente_detail["_AFFAIRE_KEY_"].isin(ok_keys)].copy()
        attente_detail = attente_detail.drop(columns=["_AFFAIRE_KEY_"], errors="ignore")

    if col_client and col_client in ok_detail.columns and col_client in attente_detail.columns:
        ok_detail["_AFFAIRE_CLIENT_AGENCE_KEY_"] = ok_detail.apply(
            lambda row: make_affaire_client_agence_key(row, col_client, col_agence),
            axis=1
        )
        attente_detail["_AFFAIRE_CLIENT_AGENCE_KEY_"] = attente_detail.apply(
            lambda row: make_affaire_client_agence_key(row, col_client, col_agence),
            axis=1
        )

        ok_client_agence_keys = set(ok_detail["_AFFAIRE_CLIENT_AGENCE_KEY_"])
        attente_detail = attente_detail[
            ~attente_detail["_AFFAIRE_CLIENT_AGENCE_KEY_"].isin(ok_client_agence_keys)
        ].copy()
        attente_detail = attente_detail.drop(columns=["_AFFAIRE_CLIENT_AGENCE_KEY_"], errors="ignore")

    return attente_detail


def sum_numeric_col(df, col):
    if df.empty or not col or col not in df.columns:
        return 0.0
    return float(pd.to_numeric(df[col], errors="coerce").fillna(0).sum())


def recompute_df_agences_attente(df_agences, df_ok, df_c, key_cols, col_client, col_agence, col_vente, col_ca_magasin, col_catalogue, col_op=None, colonnes_commerciaux=None):
    if df_agences.empty or not col_agence:
        return df_agences

    df_agences = df_agences.copy()

    for idx, row in df_agences.iterrows():
        agence = row.get("agence")
        ok_detail = df_ok[agence_mask(df_ok, agence, col_agence)].copy()
        attente_detail = df_c[agence_mask(df_c, agence, col_agence)].copy()
        attente_detail = remove_attente_already_ok(
            ok_detail,
            attente_detail,
            key_cols,
            col_client,
            col_agence,
            col_vente,
            col_ca_magasin,
            col_catalogue
        )

        ca_ok = sum_numeric_col(ok_detail, col_ca_magasin)
        ca_attente = sum_numeric_col(attente_detail, col_ca_magasin)
        nb_ok = len(ok_detail)
        nb_total = len(ok_detail) + len(attente_detail)
        detail_global = pd.concat([ok_detail, attente_detail], ignore_index=True)
        nb_opc_total = count_opc_rows(detail_global, col_op)
        ratio_opc_total = round(nb_opc_total / nb_total * 100, 2) if nb_total > 0 else 0.0
        if colonnes_commerciaux:
            bonus_malus_ok = ok_detail.apply(
                lambda ligne: calculate_bonus_malus_row(
                    ligne,
                    col_vente,
                    col_catalogue,
                    col_op,
                    colonnes_commerciaux
                ),
                axis=1
            ).sum() if not ok_detail.empty else 0.0
        else:
            bonus_malus_ok = 0.0

        df_agences.at[idx, "ca_ok"] = round(ca_ok, 2)
        df_agences.at[idx, "ca_attente"] = round(ca_attente, 2)
        df_agences.at[idx, "ca_total"] = round(ca_ok + ca_attente, 2)
        df_agences.at[idx, "ca_magasin_ok"] = round(ca_ok, 2)
        df_agences.at[idx, "bonus_malus_ok"] = round(bonus_malus_ok, 2)
        df_agences.at[idx, "nb_ok"] = nb_ok
        df_agences.at[idx, "nb_total"] = nb_total
        df_agences.at[idx, "nb_opc_total"] = nb_opc_total
        df_agences.at[idx, "ratio_opc_total_pct"] = ratio_opc_total

    return df_agences


def get_periode_start(periode):
    mois, annee = periode_to_month_year(periode)
    if not mois or not annee:
        return None
    return pd.Timestamp(year=annee, month=mois, day=1)


def collect_affaires_hors_periode(detail, col_client, col_doc, col_date, periode):
    periode_start = get_periode_start(periode)

    if detail.empty or periode_start is None or not col_date or col_date not in detail.columns:
        return []

    rows = []
    seen = set()

    for _, row in detail.iterrows():
        date_doc = pd.to_datetime(row.get(col_date), errors="coerce")

        if pd.isna(date_doc) or date_doc >= periode_start:
            continue

        client = clean_visible(row.get(col_client)) if col_client and col_client in row.index else ""
        doc = clean_visible(row.get(col_doc)) if col_doc and col_doc in row.index else ""
        client = client or "(client/affaire inconnue)"

        key = "|".join([
            normalize_key(client),
            normalize_key(doc),
            date_doc.strftime("%Y%m%d")
        ])

        if key in seen:
            continue

        seen.add(key)

        label = client
        if doc:
            label += f" (Doc: {doc})"
        label += f" - {date_doc.strftime('%d/%m/%Y')}"
        rows.append(label)

    return rows


def afficher_alerte_hors_periode(detail, col_client, col_doc, col_date, periode):
    affaires = collect_affaires_hors_periode(detail, col_client, col_doc, col_date, periode)

    if not affaires:
        return

    st.warning(
        f"⚠️ Certaines ventes ont une date de document antérieure à la période {periode}. "
        "Êtes-vous sûr que cette vente doit être comptabilisée pour cette période ?"
    )
    st.markdown("\n".join([f"- {affaire}" for affaire in affaires]))


def safe_filename(name):
    return (
        str(name)
        .replace("/", "-")
        .replace("\\", "-")
        .replace(":", "-")
        .replace("*", "-")
        .replace("?", "-")
        .replace('"', "-")
        .replace("<", "-")
        .replace(">", "-")
        .replace("|", "-")
        .strip()
    )


def pdf_text(value):
    if pd.isna(value):
        return ""
    text = str(value)
    text = text.replace("✅", "").replace("⏳", "").replace("🎯", "").replace("🌍", "")
    text = text.replace("📉", "").replace("📊", "").replace("💰", "")
    text = " ".join(text.split())
    return text


def pdf_escape(value):
    text = pdf_text(value)
    text = text.replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)")
    return text


def pdf_clip_text(text, max_chars):
    text = pdf_text(text)
    if len(text) <= max_chars:
        return text
    return text[: max(0, max_chars - 3)] + "..."


def pdf_format_value(value):
    if isinstance(value, (pd.Timestamp, datetime)):
        return value.strftime("%d/%m/%Y")
    return pdf_text(value)


@st.cache_data(show_spinner=False)
def make_simple_pdf(title, metrics, df):
    page_w, page_h = 841.89, 595.28
    margin = 24
    row_h = 16
    header_h = 18
    top_y = page_h - margin
    bottom_y = margin

    cols = list(df.columns)
    preferred_widths = {
        "Client / Référence affaire": 118,
        "Commerciaux": 122,
        "N° Document": 58,
        "Date document": 54,
        "Statut": 48,
        "Agence": 58,
        "TOTAL VENTE": 58,
        "Vente HT hors acompte": 72,
        "Total ventes avant remise": 76,
        "Remise PP": 58,
        "Remise %": 48,
        "OPC": 28,
        "Bonus / Malus": 62,
    }
    col_widths = [preferred_widths.get(col, 58) for col in cols]
    available_w = page_w - (margin * 2)
    total_w = sum(col_widths)

    if total_w > available_w:
        factor = available_w / total_w
        col_widths = [max(24, w * factor) for w in col_widths]

    pages = []
    commands = []

    def add_text(x, y, text, size=8, bold=False):
        font = "F2" if bold else "F1"
        commands.append(f"BT /{font} {size} Tf {x:.2f} {y:.2f} Td ({pdf_escape(text)}) Tj ET")

    def add_line(x1, y1, x2, y2, width=0.4):
        commands.append(f"{width:.2f} w {x1:.2f} {y1:.2f} m {x2:.2f} {y2:.2f} l S")

    def add_rect(x, y, w, h, fill=None):
        if fill:
            commands.append(f"{fill} rg {x:.2f} {y:.2f} {w:.2f} {h:.2f} re f 0 0 0 rg")
        commands.append(f"0.35 w {x:.2f} {y:.2f} {w:.2f} {h:.2f} re S")

    def new_page():
        nonlocal commands
        if commands:
            pages.append(commands)
        commands = []
        add_text(margin, top_y - 8, title, size=14, bold=True)
        y = top_y - 28
        metric_parts = [f"{k}: {v}" for k, v in metrics.items()]
        metric_lines = []
        current_line = ""

        for part in metric_parts:
            candidate = part if not current_line else f"{current_line} | {part}"
            if len(candidate) > 120 and current_line:
                metric_lines.append(current_line)
                current_line = part
            else:
                current_line = candidate

        if current_line:
            metric_lines.append(current_line)

        for metric_line in metric_lines:
            add_text(margin, y, metric_line, size=8, bold=True)
            y -= 12

        y -= 10

        x = margin
        for col, w in zip(cols, col_widths):
            add_rect(x, y - header_h + 5, w, header_h, fill="0.88 0.90 1")
            add_text(x + 2, y - 7, pdf_clip_text(col, int(w / 4.3)), size=6, bold=True)
            x += w
        return y - header_h + 5

    y = new_page()
    for _, row in df.iterrows():
        if y - row_h < bottom_y:
            y = new_page()

        x = margin
        for col, w in zip(cols, col_widths):
            add_rect(x, y - row_h, w, row_h)
            value = pdf_format_value(row.get(col, ""))
            add_text(x + 2, y - 11, pdf_clip_text(value, int(w / 3.8)), size=6)
            x += w
        y -= row_h

    pages.append(commands)

    objects = []
    objects.append(b"<< /Type /Catalog /Pages 2 0 R >>")

    page_refs = " ".join([f"{3 + i * 2} 0 R" for i in range(len(pages))])
    objects.append(f"<< /Type /Pages /Kids [{page_refs}] /Count {len(pages)} >>".encode("cp1252", errors="replace"))

    for idx, page_commands in enumerate(pages):
        page_obj_num = 3 + idx * 2
        content_obj_num = page_obj_num + 1
        page_obj = (
            f"<< /Type /Page /Parent 2 0 R /MediaBox [0 0 {page_w:.2f} {page_h:.2f}] "
            f"/Resources << /Font << /F1 << /Type /Font /Subtype /Type1 /BaseFont /Helvetica /Encoding /WinAnsiEncoding >> "
            f"/F2 << /Type /Font /Subtype /Type1 /BaseFont /Helvetica-Bold /Encoding /WinAnsiEncoding >> >> >> "
            f"/Contents {content_obj_num} 0 R >>"
        )
        objects.append(page_obj.encode("cp1252", errors="replace"))
        stream = "\n".join(page_commands).encode("cp1252", errors="replace")
        objects.append(b"<< /Length " + str(len(stream)).encode("ascii") + b" >>\nstream\n" + stream + b"\nendstream")

    pdf = bytearray(b"%PDF-1.4\n%\xe2\xe3\xcf\xd3\n")
    offsets = [0]
    for i, obj in enumerate(objects, start=1):
        offsets.append(len(pdf))
        pdf.extend(f"{i} 0 obj\n".encode("ascii"))
        pdf.extend(obj)
        pdf.extend(b"\nendobj\n")

    xref_pos = len(pdf)
    pdf.extend(f"xref\n0 {len(objects) + 1}\n".encode("ascii"))
    pdf.extend(b"0000000000 65535 f \n")
    for offset in offsets[1:]:
        pdf.extend(f"{offset:010d} 00000 n \n".encode("ascii"))
    pdf.extend(
        f"trailer\n<< /Size {len(objects) + 1} /Root 1 0 R >>\nstartxref\n{xref_pos}\n%%EOF".encode("ascii")
    )
    return bytes(pdf)


def save_pdf_export(pdf_bytes, filename):
    export_dir = DATA_DIR / "exports_pdf"
    export_dir.mkdir(parents=True, exist_ok=True)
    file_path = export_dir / filename
    file_path.write_bytes(pdf_bytes)
    return file_path.resolve()


def is_render_env():
    return bool(os.environ.get("RENDER") or os.environ.get("RENDER_SERVICE_ID"))


def build_detail_vendeur_pdf(vendeur, data, df_ok, df_c, colonnes_commerciaux, key_cols, cols, periode):
    col_client = cols["client"]
    col_doc = cols["doc"]
    col_date = cols["date"]
    col_agence = cols["agence"]
    col_vente = cols["vente"]
    col_ca_magasin = cols["ca_magasin"]
    col_catalogue = cols["catalogue"]
    col_rem = cols["rem"]
    col_op = cols["op"]

    ok_detail = df_ok[vendeur_mask(df_ok, vendeur, colonnes_commerciaux)].copy()
    attente_detail = df_c[vendeur_mask(df_c, vendeur, colonnes_commerciaux)].copy()

    attente_detail = remove_attente_already_ok(
        ok_detail,
        attente_detail,
        key_cols,
        col_client,
        col_agence,
        col_vente,
        col_ca_magasin,
        col_catalogue
    )

    ok_detail["Statut"] = "OK"
    attente_detail["Statut"] = "En attente"
    detail = pd.concat([ok_detail, attente_detail], ignore_index=True)

    if detail.empty:
        return None, None

    detail_calc = detail.copy()
    for c in [col_vente, col_ca_magasin, col_catalogue, col_rem]:
        if c and c in detail_calc.columns:
            detail_calc[c] = pd.to_numeric(detail_calc[c], errors="coerce").fillna(0)

    detail_calc["Commerciaux"] = detail_calc.apply(
        lambda row: list_commerciaux_row(row, colonnes_commerciaux),
        axis=1
    )
    detail_calc["Nombre de vendeurs"] = detail_calc.apply(
        lambda row: count_vendeurs_row(row, colonnes_commerciaux),
        axis=1
    )

    if col_catalogue and col_catalogue in detail_calc.columns and col_rem and col_rem in detail_calc.columns:
        detail_calc["Remise %"] = np.where(
            detail_calc[col_catalogue] > 0,
            detail_calc[col_rem] / detail_calc[col_catalogue] * 100,
            0
        )
    else:
        detail_calc["Remise %"] = 0

    if col_catalogue and col_catalogue in detail_calc.columns and col_vente and col_vente in detail_calc.columns:
        objectif_15_par_vendeur = (detail_calc[col_catalogue] * 0.85) / detail_calc["Nombre de vendeurs"]
        detail_calc["Bonus / Malus"] = detail_calc[col_vente] - objectif_15_par_vendeur
    else:
        detail_calc["Bonus / Malus"] = 0

    if col_op and col_op in detail_calc.columns:
        opc_mask = detail_calc.apply(lambda row: is_opc(row, col_op), axis=1)
        detail_calc.loc[opc_mask & (detail_calc["Bonus / Malus"] < 0), "Bonus / Malus"] = 0
        detail_calc["OPC"] = np.where(opc_mask, "OUI", "")
    else:
        detail_calc["OPC"] = ""

    cols_show = [
        col_client,
        "Commerciaux",
        col_doc,
        col_date,
        "Statut",
        col_agence,
        col_vente,
        col_ca_magasin,
        "Remise %",
        "OPC",
        "Bonus / Malus"
    ]
    cols_show = list(dict.fromkeys([c for c in cols_show if c and c in detail_calc.columns]))
    detail_affichage = detail_calc[cols_show].copy()

    rename_cols = {}
    if col_client:
        rename_cols[col_client] = "Client / Référence affaire"
    if col_doc:
        rename_cols[col_doc] = "N° Document"
    if col_date:
        rename_cols[col_date] = "Date document"
    if col_agence:
        rename_cols[col_agence] = "Agence"
    if col_vente:
        rename_cols[col_vente] = "TOTAL VENTE"
    if col_ca_magasin:
        rename_cols[col_ca_magasin] = "Vente HT hors acompte"
    detail_affichage = detail_affichage.rename(columns=rename_cols)

    for c in ["TOTAL VENTE", "Vente HT hors acompte", "Remise %", "Bonus / Malus"]:
        if c in detail_affichage.columns:
            detail_affichage[c] = pd.to_numeric(detail_affichage[c], errors="coerce").fillna(0).round(2)

    if "Date document" in detail_affichage.columns:
        detail_affichage["Date document"] = pd.to_datetime(
            detail_affichage["Date document"],
            errors="coerce"
        ).dt.strftime("%d/%m/%Y").fillna("")

    metrics = {
        "CA OK": f"{to_float(data.get('ca_ok', 0)):,.2f} EUR",
        "CA attente": f"{to_float(data.get('ca_attente', 0)):,.2f} EUR",
        "CA Total": f"{to_float(data.get('ca_total', 0)):,.2f} EUR",
        "Commission": f"{to_float(data.get('commission_eur', 0)):,.2f} EUR",
        "Remise commission hors OPC": f"{to_float(data.get('remise_hors_opc_pct', 0)):,.2f} %",
        "Base commission": f"{to_float(data.get('base_commission_pct', 0)):,.2f} %",
        "Points perdus": f"{to_float(data.get('points_perdus', 0)):,.0f}",
        "Commission definitive": f"{to_float(data.get('commission_pct', 0)):,.2f} %",
        "Remise OK hors OPC": f"{to_float(data.get('remise_ok_pct', 0)):,.2f} %",
        "Bonus / Malus OK": f"{to_float(data.get('bonus_malus_ok', 0)):,.2f} EUR",
    }

    return detail_affichage, metrics


def build_detail_agence_pdf(agence, data, df_ok, df_c, key_cols, cols, periode):
    col_client = cols["client"]
    col_doc = cols["doc"]
    col_date = cols["date"]
    col_agence = cols["agence"]
    col_vente = cols["vente"]
    col_ca_magasin = cols["ca_magasin"]
    col_catalogue = cols["catalogue"]
    col_rem = cols["rem"]
    col_op = cols["op"]
    colonnes_commerciaux = cols["commerciaux"]

    ok_detail = df_ok[agence_mask(df_ok, agence, col_agence)].copy()
    attente_detail = df_c[agence_mask(df_c, agence, col_agence)].copy()

    attente_detail = remove_attente_already_ok(
        ok_detail,
        attente_detail,
        key_cols,
        col_client,
        col_agence,
        col_vente,
        col_ca_magasin,
        col_catalogue
    )

    ok_detail["Statut"] = "OK"
    attente_detail["Statut"] = "En attente"
    detail = pd.concat([ok_detail, attente_detail], ignore_index=True)

    if detail.empty:
        return None, None

    detail["Commerciaux"] = detail.apply(
        lambda row: list_commerciaux_row(row, colonnes_commerciaux),
        axis=1
    )

    cols_show = [
        col_client,
        "Commerciaux",
        col_doc,
        col_date,
        "Statut",
        col_agence,
        col_ca_magasin,
        col_vente,
        col_catalogue,
        col_rem,
        col_op
    ]
    cols_show = list(dict.fromkeys([c for c in cols_show if c and c in detail.columns]))
    detail_affichage = detail[cols_show].copy()

    for c in [col_ca_magasin, col_vente, col_catalogue, col_rem]:
        if c and c in detail_affichage.columns:
            detail_affichage[c] = pd.to_numeric(detail_affichage[c], errors="coerce").fillna(0).round(2)

    if col_date and col_date in detail_affichage.columns:
        detail_affichage[col_date] = pd.to_datetime(
            detail_affichage[col_date],
            errors="coerce"
        ).dt.strftime("%d/%m/%Y").fillna("")

    metrics = {
        "CA OK agence": f"{to_float(data.get('ca_ok', 0)):,.2f} EUR",
        "CA attente agence": f"{to_float(data.get('ca_attente', 0)):,.2f} EUR",
        "CA Total agence": f"{to_float(data.get('ca_total', 0)):,.2f} EUR",
        "Bonus / Malus OK": f"{to_float(data.get('bonus_malus_ok', 0)):,.2f} EUR",
        "Remise moyenne": f"{to_float(data.get('remise_pct', 0)):,.2f} %",
    }

    return detail_affichage, metrics


def make_pdf_zip(items):
    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        for filename, pdf_bytes in items:
            zf.writestr(filename, pdf_bytes)
    buffer.seek(0)
    return buffer.getvalue()


def update_evp_workbook(evp_file, df_vendeurs, df_directeurs, periode):
    target_sheet = sheet_name_evp_for_next_month(periode)
    if not target_sheet:
        return None, f"Période non reconnue : {periode}", []

    wb = load_workbook(evp_file)
    sheet_lookup = {strip_accents(normalize_key(ws.title)): ws for ws in wb.worksheets}
    ws = sheet_lookup.get(strip_accents(normalize_key(target_sheet)))

    if ws is None:
        return None, f"Onglet EVP introuvable : {target_sheet}", []

    row_by_name = {}
    for row_idx in range(1, ws.max_row + 1):
        nom = clean_visible(ws.cell(row=row_idx, column=2).value)
        key = resolve_nom_evp(nom)
        if key and key not in row_by_name:
            row_by_name[key] = row_idx

    absents = []

    if df_vendeurs is not None and not df_vendeurs.empty:
        for _, vendeur in df_vendeurs.iterrows():
            nom = vendeur.get("Commercial", "")
            if is_excluded_from_evp(nom):
                continue

            key = resolve_nom_evp(nom)
            row_idx = row_by_name.get(key)

            if not row_idx:
                absents.append(clean_visible(nom) or key)
                continue

            ca_ok = to_float(vendeur.get("ca_ok", 0))
            commission_pct = to_float(vendeur.get("commission_pct", 0))
            commission_eur = to_float(vendeur.get("commission_eur", 0))
            points_perdus = to_float(vendeur.get("points_perdus", 0))

            ws.cell(row=row_idx, column=5).value = ca_ok
            ws.cell(row=row_idx, column=5).number_format = '#,##0.00 €'
            ws.cell(row=row_idx, column=6).value = commission_pct / 100
            ws.cell(row=row_idx, column=6).number_format = '0.00%'

            if commission_pct == 0:
                ws.cell(row=row_idx, column=7).value = None
                ws.cell(row=row_idx, column=8).value = None
            else:
                montant_cell = ws.cell(row=row_idx, column=7)
                if not isinstance(montant_cell.value, str) or not montant_cell.value.startswith("="):
                    montant_cell.value = commission_eur
                    montant_cell.number_format = '#,##0.00 €'
                ws.cell(row=row_idx, column=7).number_format = '#,##0.00 €'
                ws.cell(row=row_idx, column=8).value = points_perdus / 100
                ws.cell(row=row_idx, column=8).number_format = '0.00%'

            ws.cell(row=row_idx, column=14).value = None

    if df_directeurs is not None and not df_directeurs.empty:
        for _, directeur in df_directeurs.iterrows():
            nom = directeur.get("directeur", "")
            key = resolve_nom_evp(nom)
            row_idx = row_by_name.get(key)

            if not row_idx:
                absents.append(clean_visible(nom) or key)
                continue

            montant = to_float(directeur.get("commission_magasin_eur", 0))
            if montant > 0:
                ws.cell(row=row_idx, column=10).value = montant
                ws.cell(row=row_idx, column=10).number_format = '#,##0.00 €'
            else:
                ws.cell(row=row_idx, column=10).value = None

    try:
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
    except Exception:
        pass

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue(), None, sorted(set(absents))


def save_periode(periode, data):
    file_path = HISTORIQUE_DIR / f"{safe_filename(periode)}.pkl"
    with open(file_path, "wb") as f:
        pickle.dump(data, f)
    load_periode_cached.clear()
    load_all_historique_cached.clear()


@st.cache_data(show_spinner=False)
def load_periode_cached(periode, file_mtime_ns):
    file_path = HISTORIQUE_DIR / f"{safe_filename(periode)}.pkl"
    with open(file_path, "rb") as f:
        return pickle.load(f)


def load_periode(periode):
    file_path = HISTORIQUE_DIR / f"{safe_filename(periode)}.pkl"
    if not file_path.exists():
        return None
    return load_periode_cached(periode, file_path.stat().st_mtime_ns)


def load_periode_preserve_ui(periode):
    saved_data = load_periode(periode)
    if not saved_data:
        return False

    ui_keys = [
        "active_page_admin",
        "active_page_directeur",
        "active_page_vendeur",
        "vendeur_select",
        "agence_select",
    ]
    preserved_ui = {
        key: st.session_state.get(key)
        for key in ui_keys
        if key in st.session_state
    }

    st.session_state.update(saved_data)

    for key, value in preserved_ui.items():
        if value is not None:
            st.session_state[key] = value

    return True


def delete_periode(periode):
    file_path = HISTORIQUE_DIR / f"{safe_filename(periode)}.pkl"
    if file_path.exists():
        file_path.unlink()
        load_periode_cached.clear()
        load_all_historique_cached.clear()
        return True
    return False


def list_periodes():
    return sorted([f.stem for f in HISTORIQUE_DIR.glob("*.pkl")])


# ====================== FONCTIONS PÉRIODES / ANNUEL ======================

MOIS_FR = {
    "JANVIER": 1,
    "FEVRIER": 2,
    "FÉVRIER": 2,
    "MARS": 3,
    "AVRIL": 4,
    "MAI": 5,
    "JUIN": 6,
    "JUILLET": 7,
    "AOUT": 8,
    "AOÛT": 8,
    "SEPTEMBRE": 9,
    "OCTOBRE": 10,
    "NOVEMBRE": 11,
    "DECEMBRE": 12,
    "DÉCEMBRE": 12,
}


def periode_to_month_year(periode):
    if not periode:
        return None, None

    parts = str(periode).strip().upper().split()

    if len(parts) < 2:
        return None, None

    mois = MOIS_FR.get(parts[0])
    annee = None

    for p in parts[1:]:
        if p.isdigit() and len(p) == 4:
            annee = int(p)
            break

    return mois, annee


def periode_sort_key(periode):
    mois, annee = periode_to_month_year(periode)
    return (annee or 0, mois or 0)


def sheet_name_evp_for_next_month(periode):
    mois, annee = periode_to_month_year(periode)
    if not mois or not annee:
        return None

    next_month = mois + 1
    next_year = annee

    if next_month > 12:
        next_month = 1
        next_year += 1

    mois_names = {
        1: "JANVIER",
        2: "FEVRIER",
        3: "MARS",
        4: "AVRIL",
        5: "MAI",
        6: "JUIN",
        7: "JUILLET",
        8: "AOUT",
        9: "SEPTEMBRE",
        10: "OCTOBRE",
        11: "NOVEMBRE",
        12: "DECEMBRE",
    }
    return f"{mois_names[next_month]} {next_year}"


def is_periode_comptable_annuelle(periode, annee_selectionnee, use_m2=True):
    mois, annee = periode_to_month_year(periode)

    if not mois or not annee:
        return False

    if annee != annee_selectionnee:
        return False

    if not use_m2:
        return True

    today = datetime.today()

    if annee < today.year:
        return True

    if annee > today.year:
        return False

    return mois <= today.month - 2


def historique_signature():
    return tuple(
        sorted((f.name, f.stat().st_mtime_ns, f.stat().st_size) for f in HISTORIQUE_DIR.glob("*.pkl"))
    )


@st.cache_data(show_spinner=False)
def load_all_historique_cached(signature):
    rows_vendeurs = []
    rows_agences = []
    rows_directeurs = []

    for periode in list_periodes():
        data = load_periode(periode)

        if not data:
            continue

        mois, annee = periode_to_month_year(periode)

        if "df_vendeurs" in data and isinstance(data["df_vendeurs"], pd.DataFrame):
            dfv = ensure_bonus_malus_columns(data["df_vendeurs"].copy())
            dfv = ensure_remise_columns(dfv)
            dfv["periode"] = periode
            dfv["mois_num"] = mois
            dfv["annee"] = annee
            rows_vendeurs.append(dfv)

        if "df_agences" in data and isinstance(data["df_agences"], pd.DataFrame):
            dfa = data["df_agences"].copy()
            dfa["periode"] = periode
            dfa["mois_num"] = mois
            dfa["annee"] = annee
            rows_agences.append(dfa)

        if "df_directeurs" in data and isinstance(data["df_directeurs"], pd.DataFrame):
            dfd = data["df_directeurs"].copy()
            dfd["periode"] = periode
            dfd["mois_num"] = mois
            dfd["annee"] = annee
            rows_directeurs.append(dfd)

    df_all_vendeurs = pd.concat(rows_vendeurs, ignore_index=True) if rows_vendeurs else pd.DataFrame()
    df_all_agences = pd.concat(rows_agences, ignore_index=True) if rows_agences else pd.DataFrame()
    df_all_directeurs = pd.concat(rows_directeurs, ignore_index=True) if rows_directeurs else pd.DataFrame()

    return df_all_vendeurs, df_all_agences, df_all_directeurs


def load_all_historique():
    return load_all_historique_cached(historique_signature())


# ====================== FORMAT TABLEAUX ======================

def format_df_vendeurs(df):
    if df.empty:
        return df
    technical_cols = [
        "ca_commissionnable_directeur_ok",
        "ca_commissionnable_directeur_total",
        "nb_remise_plus_30_ok",
        "nb_remise_plus_30_total",
    ]
    df = df.drop(columns=technical_cols, errors="ignore")
    return df.rename(columns={
        "ca_ok": "CA OK",
        "ca_attente": "CA en attente",
        "ca_total": "CA Total",
        "remise_hors_opc_pct": "Remise moy. % hors OPC",
        "remise_hors_opc_global_pct": "Remise globale % hors OPC",
        "remise_ok_pct": "Remise OK hors OPC",
        "remise_global_pct": "Remise Global hors OPC",
        "base_commission_pct": "Base commission %",
        "points_perdus": "Points perdus",
        "commission_pct": "% Commission",
        "commission_eur": "Commission €",
        "bonus_malus_ok": "Bonus / Malus OK",
        "bonus_malus_global": "Bonus / Malus Global",
        "nb_ok": "Nb affaires OK",
        "nb_opc_total": "Nb OPC total",
        "ratio_opc_total_pct": "Ratio OPC total %",
    })


def format_df_agences(df):
    if df.empty:
        return df
    return df.rename(columns={
        "agence": "Agence",
        "ca_ok": "CA OK",
        "ca_attente": "CA en attente",
        "ca_total": "CA Total",
        "ca_magasin_ok": "CA magasin OK",
        "bonus_malus_ok": "Bonus / Malus OK",
        "remise_pct": "Remise moyenne %",
        "nb_ok": "Nb affaires OK",
        "nb_total": "Nb affaires total",
        "nb_opc_total": "Nb OPC total",
        "ratio_opc_total_pct": "Ratio OPC total %",
    })


def format_df_directeurs(df):
    if df.empty:
        return df
    return df.rename(columns={
        "directeur": "Directeur",
        "agence": "Agence",
        "ca_magasin_ok": "CA magasin OK",
        "un_pourcent_ca": "1% CA",
        "prime_palier": "Prime palier",
        "commission_magasin_eur": "Commission magasin €",
    })


# ====================== LOGIN ======================

if "logged_in" not in st.session_state:
    st.session_state.logged_in = False

if not st.session_state.logged_in:
    logo_path = Path("logo.png")
    if logo_path.exists():
        logo_b64 = base64.b64encode(logo_path.read_bytes()).decode("utf-8")
        logo_html = f'<img src="data:image/png;base64,{logo_b64}" alt="EcoHabitat">'
    else:
        logo_html = "<div class='login-logo-fallback'>🏠</div>"

    login_video_html = ""
    for video_name, mime_type in [
        ("login_background.mp4", "video/mp4"),
        ("login_background.webm", "video/webm"),
    ]:
        video_path = Path(video_name)
        if video_path.exists():
            video_b64 = base64.b64encode(video_path.read_bytes()).decode("utf-8")
            login_video_html = (
                f'<video class="login-video-bg" autoplay muted loop playsinline>'
                f'<source src="data:{mime_type};base64,{video_b64}" type="{mime_type}">'
                f'</video><div class="login-video-overlay"></div>'
            )
            break

    st.markdown(
        f"""
        <style>
        .login-video-bg {{
            position: fixed;
            inset: 0;
            width: 100vw;
            height: 100vh;
            object-fit: cover;
            z-index: 0;
            opacity: 0.24;
            pointer-events: none;
        }}

        .login-video-overlay {{
            position: fixed;
            inset: 0;
            background: rgba(246, 248, 244, 0.48);
            backdrop-filter: blur(0.5px);
            z-index: 1;
            pointer-events: none;
        }}

        [data-testid="stAppViewContainer"] > .main,
        [data-testid="stHeader"] {{
            position: relative;
            z-index: 2;
        }}

        .login-brand {{
            display: flex;
            flex-direction: column;
            align-items: center;
            justify-content: center;
            width: 100%;
            margin: 56px 0 22px;
            text-align: center;
            position: relative;
            z-index: 4;
            opacity: 1 !important;
        }}

        .login-brand img {{
            width: 150px;
            display: block;
            margin: 0 auto 24px;
            opacity: 1 !important;
            filter: drop-shadow(0 8px 16px rgba(18, 38, 58, 0.16));
        }}

        .login-brand h1 {{
            margin: 0;
            font-size: 40px;
            font-weight: 800;
            text-align: center;
            color: #12263A !important;
            opacity: 1 !important;
            text-shadow: 0 2px 8px rgba(255, 255, 255, 0.85);
        }}

        [data-testid="stTextInput"],
        [data-testid="stButton"] {{
            position: relative;
            z-index: 4;
        }}

        .login-logo-fallback {{
            font-size: 78px;
            line-height: 1;
            margin-bottom: 24px;
        }}
        </style>
        {login_video_html}
        <div class="login-brand">
            {logo_html}
            <h1>🔐 Connexion - ECOHABITAT</h1>
        </div>
        """,
        unsafe_allow_html=True
    )

    username = st.text_input("Identifiant")
    password = st.text_input("Mot de passe", type="password")
    totp_code = st.text_input("Code Authenticator", max_chars=6, help="À renseigner uniquement si la 2FA est activée.")

    pending_2fa_user = st.session_state.get("pending_2fa_setup_user")
    if pending_2fa_user:
        users = load_users()
        pending_user = users.get(pending_2fa_user)
        if pending_user and pending_user.get("totp_enabled") and not pending_user.get("totp_confirmed"):
            pending_secret = pending_user.get("totp_secret") or generate_totp_secret()
            pending_user["totp_secret"] = pending_secret
            users[pending_2fa_user] = pending_user
            save_users(users)

            st.info("Première connexion 2FA : scanne ce QR code avec Google Authenticator ou Microsoft Authenticator, puis saisis le code généré.")
            st.image(qr_code_url(totp_uri(pending_2fa_user, pending_secret)), width=220)
            st.code(pending_secret, language=None)
            first_totp_code = st.text_input("Code après scan", max_chars=6, key="first_totp_code")

            if st.button("Activer Authenticator et se connecter", type="primary"):
                if verify_totp(pending_secret, first_totp_code):
                    pending_user["totp_confirmed"] = True
                    users[pending_2fa_user] = pending_user
                    save_users(users)
                    reset_login_security_state()
                    st.session_state.pop("pending_2fa_setup_user", None)
                    st.session_state.logged_in = True
                    st.session_state.user = pending_user
                    st.session_state.username = pending_2fa_user
                    st.session_state.pop(f"sidebar_auto_collapsed_{pending_user.get('role')}", None)
                    st.rerun()
                else:
                    st.error("Code Authenticator incorrect.")
        else:
            st.session_state.pop("pending_2fa_setup_user", None)

    if st.button("Se connecter", type="primary"):
        remaining_lock = get_login_lock_remaining()

        if remaining_lock > 0:
            st.error(f"Trop d'essais incorrects. Réessaie dans {remaining_lock // 60 + 1} min.")
        else:
            users = load_users()
            username_key = username.lower().strip()
            user = users.get(username_key)
            stored_password = user.get("password_hash") or user.get("password") if user else ""

            if user and verify_password(password, stored_password):
                if user.get("totp_enabled") and not user.get("totp_confirmed"):
                    if not user.get("totp_secret"):
                        user["totp_secret"] = generate_totp_secret()
                    users[username_key] = user
                    save_users(users)
                    st.session_state.pending_2fa_setup_user = username_key
                    st.rerun()

                if user.get("totp_enabled"):
                    if not verify_totp(user.get("totp_secret"), totp_code):
                        register_failed_login()
                        st.error("Code Authenticator incorrect.")
                        st.stop()

                if not password_is_hashed(user):
                    user["password_hash"] = hash_password(password)
                    user.pop("password", None)
                    users[username_key] = user
                    save_users(users)

                reset_login_security_state()
                st.session_state.logged_in = True
                st.session_state.user = user
                st.session_state.username = username_key
                st.session_state.pop(f"sidebar_auto_collapsed_{user.get('role')}", None)
                st.rerun()
            else:
                register_failed_login()
                remaining_attempts = MAX_LOGIN_ATTEMPTS - st.session_state.get("login_attempts", 0)
                if remaining_attempts > 0:
                    st.error(f"Identifiant ou mot de passe incorrect. Essais restants : {remaining_attempts}.")
                else:
                    st.error("Trop d'essais incorrects. Connexion bloquée temporairement.")

    st.stop()

user = st.session_state.user
role = user["role"]

if role in ["admin", "vendeur", "directeur_agence"] and not st.session_state.get(f"sidebar_auto_collapsed_{role}"):
    components.html(
        """
        <script>
        let attempts = 0;
        const closeSidebar = () => {
            attempts += 1;
            const doc = window.parent.document;
            const sidebar = doc.querySelector('[data-testid="stSidebar"]');
            const collapseButton =
                doc.querySelector('[data-testid="stSidebarCollapseButton"]') ||
                doc.querySelector('button[aria-label="Close sidebar"]') ||
                doc.querySelector('button[title="Close sidebar"]');
            const isOpen = sidebar && sidebar.getBoundingClientRect().width > 80;
            if (isOpen && collapseButton) {
                collapseButton.click();
                return;
            }
            if (attempts < 20) {
                setTimeout(closeSidebar, 150);
            }
        };
        setTimeout(closeSidebar, 150);
        </script>
        """,
        height=0,
        width=0
    )
    st.session_state[f"sidebar_auto_collapsed_{role}"] = True


col1, col2 = st.columns([1, 6])

with col1:
    try:
        st.image("logo.png", width=90)
    except Exception:
        st.markdown("🏠")

with col2:
    st.markdown('<div class="eco-title">gesCom EcoHabitat</div>', unsafe_allow_html=True)
    st.markdown('<div class="eco-subtitle">L’excellence au service de votre habitat</div>', unsafe_allow_html=True)

st.sidebar.success(f"Connecté : {user['nom']} ({role})")

if st.sidebar.button("🚪 Déconnexion"):
    for k in [
        "logged_in", "user", "username",
        "sidebar_auto_collapsed_admin",
        "sidebar_auto_collapsed_vendeur",
        "sidebar_auto_collapsed_directeur_agence",
    ]:
        if k in st.session_state:
            del st.session_state[k]
    st.rerun()


# ====================== SIDEBAR HISTORIQUE ======================

st.sidebar.markdown("---")
st.sidebar.subheader("📂 Historique")

periodes_dispo = sorted(list_periodes(), key=periode_sort_key)

periode_load = st.sidebar.selectbox(
    "Charger une période sauvegardée",
    [""] + periodes_dispo
)

if st.sidebar.button("📥 Charger la période"):
    if not periode_load:
        st.warning("Sélectionne une période.")
    else:
        if load_periode_preserve_ui(periode_load):
            st.success(f"✅ Période {periode_load} chargée.")
        else:
            st.error("❌ Impossible de charger cette période.")

if role == "admin" and periodes_dispo:
    with st.sidebar.expander("🗑️ Supprimer une période", expanded=False):
        periode_delete = st.selectbox(
            "Période à supprimer",
            [""] + periodes_dispo,
            key="periode_delete_select"
        )
        confirm_delete = st.checkbox(
            "Je confirme la suppression",
            key="periode_delete_confirm"
        )

        if st.button(
            "Supprimer la période",
            disabled=not periode_delete or not confirm_delete,
            key="periode_delete_button"
        ):
            if delete_periode(periode_delete):
                if st.session_state.get("periode") == periode_delete:
                    keys_to_clear = [
                        "df_vendeurs", "df_agences", "df_directeurs",
                        "df_ok", "df_c", "periode"
                    ]
                    for key in keys_to_clear:
                        st.session_state.pop(key, None)
                st.success(f"✅ Période {periode_delete} supprimée.")
                st.rerun()
            else:
                st.error("❌ Impossible de supprimer cette période.")


# ====================== IMPORT ADMIN ======================

if role == "admin":

    st.sidebar.markdown("---")
    with st.sidebar.expander("📤 Import ProDevis", expanded=False):

        f_confirm = st.file_uploader("Fichier CONFIRM", type=["xlsx"], key="upload_confirm")
        f_ok = st.file_uploader("Fichier BONLIVR", type=["xlsx"], key="upload_ok")

        periode = st.text_input("📅 Période", value="Avril 2026", key="periode_input")

        lancer_traitement = st.button("🚀 Lancer le traitement", type="primary")

    if lancer_traitement:

        if not f_confirm or not f_ok:
            st.error("❌ Charge les deux fichiers : CONFIRM et BONLIVR.")
            st.stop()

        df_confirm = pd.read_excel(f_confirm, skiprows=28, header=0)
        df_ok = pd.read_excel(f_ok, skiprows=28, header=0)

        for df in [df_confirm, df_ok]:
            df.columns = [str(col).replace("\n", " ").strip() for col in df.columns]

        df_confirm["Statut"] = "⏳ En attente"
        df_ok["Statut"] = "✅ OK"

        col_client = get_col_by_excel_position(df_confirm, 1)
        col_com1 = get_col_by_excel_position(df_confirm, 2)
        col_com2 = get_col_by_excel_position(df_confirm, 3)
        col_com3 = get_col_by_excel_position(df_confirm, 4)
        col_date = get_col_by_excel_position(df_confirm, 6)
        col_doc = get_col_by_excel_position(df_confirm, 7)
        col_op = get_col_by_excel_position(df_confirm, 8)
        col_ca_magasin = get_col_by_excel_position(df_confirm, 9)
        col_vente = get_col_by_excel_position(df_confirm, 17)
        col_rem = get_col_by_excel_position(df_confirm, 24)
        col_catalogue = get_col_by_excel_position(df_confirm, 25)
        col_agence = get_col_by_excel_position(df_confirm, 50)

        col_vente = col_vente or find_col(df_confirm, includes=["TOTAL VENTE"])
        col_catalogue = col_catalogue or find_col(df_confirm, includes=["TOTAL VENTES AVANT REMISE"])
        col_rem = col_rem or find_col(df_confirm, includes=["REMISE PP"])
        col_op = col_op or find_col(df_confirm, includes=["OPERATION COMMERCIALE"])
        col_agence = col_agence or find_col(df_confirm, includes=["AGENCE"])
        col_ca_magasin = col_ca_magasin or find_col(df_confirm, includes=["VENTE HT"])

        if not col_vente:
            st.error("❌ Colonne vendeur introuvable : Q / TOTAL VENTE.")
            st.stop()

        if not col_ca_magasin:
            st.error("❌ Colonne agence introuvable : I / Vente HT hors acompte.")
            st.stop()

        if not col_catalogue:
            st.error("❌ Colonne catalogue introuvable : Y / TOTAL VENTES AVANT REMISE.")
            st.stop()

        if not col_rem:
            st.error("❌ Colonne remise PP introuvable : X / REMISE PP.")
            st.stop()

        if not col_agence:
            st.error("❌ Colonne agence introuvable : AX.")
            st.stop()

        colonnes_commerciaux = [col_com1, col_com2, col_com3]

        key_cols = []
        for c in [col_client, col_date, col_doc]:
            if c and c in df_confirm.columns and c in df_ok.columns and c not in key_cols:
                key_cols.append(c)

        # ====================== VENDEURS ======================

        vendors = {}

        for _, row in df_confirm.iterrows():

            vente = pd.to_numeric(row.get(col_vente), errors="coerce")
            vente = float(vente) if pd.notna(vente) else 0.0

            if vente <= 0:
                continue

            bonus_malus = calculate_bonus_malus_row(row, col_vente, col_catalogue, col_op, colonnes_commerciaux)

            for col in colonnes_commerciaux:
                if not col or col not in df_confirm.columns:
                    continue

                nom = clean_visible(row.get(col))

                if not nom:
                    continue

                k = normalize_key(nom)

                if k not in vendors:
                    vendors[k] = {
                        "nom": nom,
                        "total": 0.0,
                        "ok": 0.0,
                        "rem_hors_opc": 0.0,
                        "catalogue_hors_opc": 0.0,
                        "bonus_malus_ok": 0.0,
                        "bonus_malus_global": 0.0
                    }

                vendors[k]["total"] += vente
                vendors[k]["bonus_malus_global"] += bonus_malus

        for _, row in df_ok.iterrows():

            vente = pd.to_numeric(row.get(col_vente), errors="coerce")
            vente = float(vente) if pd.notna(vente) else 0.0

            if vente <= 0:
                continue

            rem = pd.to_numeric(row.get(col_rem), errors="coerce")
            rem = float(rem) if pd.notna(rem) else 0.0

            catalogue = pd.to_numeric(row.get(col_catalogue), errors="coerce")
            catalogue = float(catalogue) if pd.notna(catalogue) else 0.0

            opc = is_opc(row, col_op)
            bonus_malus = calculate_bonus_malus_row(row, col_vente, col_catalogue, col_op, colonnes_commerciaux)

            for col in colonnes_commerciaux:
                if not col or col not in df_ok.columns:
                    continue

                nom = clean_visible(row.get(col))

                if not nom:
                    continue

                k = normalize_key(nom)

                if k not in vendors:
                    vendors[k] = {
                        "nom": nom,
                        "total": 0.0,
                        "ok": 0.0,
                        "rem_hors_opc": 0.0,
                        "catalogue_hors_opc": 0.0,
                        "bonus_malus_ok": 0.0,
                        "bonus_malus_global": 0.0
                    }

                vendors[k]["ok"] += vente
                vendors[k]["bonus_malus_ok"] += bonus_malus

                if not opc:
                    vendors[k]["rem_hors_opc"] += rem
                    vendors[k]["catalogue_hors_opc"] += catalogue

        vendeur_results = []

        for _, v in vendors.items():

            if is_excluded_from_evp(v["nom"]):
                continue

            remise_commission = (
                v["rem_hors_opc"] / v["catalogue_hors_opc"] * 100
                if v["catalogue_hors_opc"] > 0
                else 0
            )

            ca_attente = max(0, v["total"] - v["ok"])
            base_comm, points, comm_def, euro = calculate_commission(v["ok"], remise_commission)

            vendeur_results.append({
                "Commercial": v["nom"],
                "ca_ok": round(v["ok"], 2),
                "ca_attente": round(ca_attente, 2),
                "ca_total": round(v["total"], 2),
                "remise_hors_opc_pct": round(remise_commission, 2),
                "base_commission_pct": base_comm,
                "points_perdus": points,
                "commission_pct": comm_def,
                "commission_eur": euro,
                "bonus_malus_ok": round(v["bonus_malus_ok"], 2),
                "bonus_malus_global": round(v["bonus_malus_global"], 2)
            })

        df_vendeurs = recompute_df_vendeurs_indicators(
            pd.DataFrame(vendeur_results),
            df_ok,
            df_confirm,
            col_vente,
            col_catalogue,
            col_rem,
            col_op,
            colonnes_commerciaux,
            key_cols,
            col_client,
            col_agence,
            col_ca_magasin
        )

        # ====================== AGENCES ======================

        agences = {}

        for _, row in df_confirm.iterrows():

            agence = clean_visible(row.get(col_agence))
            if not agence:
                continue

            k = normalize_key(agence)

            ca_agence = pd.to_numeric(row.get(col_ca_magasin), errors="coerce")
            ca_agence = float(ca_agence) if pd.notna(ca_agence) else 0.0

            if k not in agences:
                agences[k] = {
                    "agence": agence,
                    "total": 0.0,
                    "ok": 0.0,
                    "rem": 0.0,
                    "catalogue": 0.0,
                    "nb_confirm": 0,
                    "nb_ok": 0
                }

            agences[k]["total"] += ca_agence
            agences[k]["nb_confirm"] += 1

        for _, row in df_ok.iterrows():

            agence = clean_visible(row.get(col_agence))
            if not agence:
                continue

            k = normalize_key(agence)

            ca_agence = pd.to_numeric(row.get(col_ca_magasin), errors="coerce")
            ca_agence = float(ca_agence) if pd.notna(ca_agence) else 0.0

            rem = pd.to_numeric(row.get(col_rem), errors="coerce")
            rem = float(rem) if pd.notna(rem) else 0.0

            catalogue = pd.to_numeric(row.get(col_catalogue), errors="coerce")
            catalogue = float(catalogue) if pd.notna(catalogue) else 0.0

            if k not in agences:
                agences[k] = {
                    "agence": agence,
                    "total": 0.0,
                    "ok": 0.0,
                    "rem": 0.0,
                    "catalogue": 0.0,
                    "nb_confirm": 0,
                    "nb_ok": 0
                }

            agences[k]["ok"] += ca_agence
            agences[k]["rem"] += rem
            agences[k]["catalogue"] += catalogue
            agences[k]["nb_ok"] += 1

        agence_results = []

        for _, a in agences.items():

            remise_agence = (a["rem"] / a["catalogue"] * 100) if a["catalogue"] > 0 else 0
            ca_attente = max(0, a["total"] - a["ok"])

            agence_results.append({
                "agence": a["agence"],
                "ca_ok": round(a["ok"], 2),
                "ca_attente": round(ca_attente, 2),
                "ca_total": round(a["total"], 2),
                "ca_magasin_ok": round(a["ok"], 2),
                "remise_pct": round(remise_agence, 2),
                "nb_ok": a["nb_ok"],
                "nb_total": a["nb_confirm"]
            })

        df_agences = pd.DataFrame(agence_results)
        df_agences = recompute_df_agences_attente(
            df_agences,
            df_ok,
            df_confirm,
            key_cols,
            col_client,
            col_agence,
            col_vente,
            col_ca_magasin,
            col_catalogue
        )

        # ====================== DIRECTEURS ======================

        rules_directeurs = [
            {"directeur": "VUE JONATHAN", "agence": "BOURG ACHARD"},
            {"directeur": "AYACHE ADEL", "agence": "MAROMME"},
            {"directeur": "EL GHAZOUANI NAHIM", "agence": "HARFLEUR"},
        ]

        directeur_results = []

        ca_by_agence = {
            normalize_key(row["agence"]): row["ca_magasin_ok"]
            for _, row in df_agences.iterrows()
        } if not df_agences.empty else {}

        for rule in rules_directeurs:
            ag_key = normalize_key(rule["agence"])
            ca = ca_by_agence.get(ag_key, 0.0)
            prime = prime_magasin(ca)
            comm_magasin = round((ca * 0.01) + prime, 2) if ca > 0 else 0.0

            directeur_results.append({
                "directeur": rule["directeur"],
                "agence": rule["agence"],
                "ca_magasin_ok": round(ca, 2),
                "un_pourcent_ca": round(ca * 0.01, 2),
                "prime_palier": prime,
                "commission_magasin_eur": comm_magasin
            })

        df_directeurs = pd.DataFrame(directeur_results)

        if df_vendeurs.empty:
            st.error("❌ Aucun vendeur trouvé.")
            st.stop()

        data_to_save = {
            "df_vendeurs": df_vendeurs,
            "df_agences": df_agences,
            "df_directeurs": df_directeurs,
            "df_ok": df_ok,
            "df_c": df_confirm,
            "col_client": col_client,
            "col_doc": col_doc,
            "col_date": col_date,
            "col_op": col_op,
            "col_ca_magasin": col_ca_magasin,
            "col_vente": col_vente,
            "col_rem": col_rem,
            "col_catalogue": col_catalogue,
            "col_agence": col_agence,
            "col_com1": col_com1,
            "col_com2": col_com2,
            "col_com3": col_com3,
            "key_cols": key_cols,
            "periode": periode
        }

        st.session_state.update(data_to_save)
        save_periode(periode, data_to_save)

        st.success(f"✅ {periode} chargé et sauvegardé avec succès !")


# ====================== OUTILS SIDEBAR ======================

st.sidebar.markdown("---")
with st.sidebar.expander("🧰 Outils", expanded=False):
    use_m2_rule = st.checkbox(
        "Règle M-2",
        value=True,
        help="Filtre l'analyse annuelle sur les mois comptablement finalisés."
    )

    if not use_m2_rule:
        st.caption("Mode test : M-2 désactivé")

    if role == "admin" and st.session_state.get("df_vendeurs") is not None:
        st.divider()
        st.caption("📊 Mise à jour EVP mensuelle")

        periode_tools = st.session_state.get("periode", "Mois inconnu")
        df_vendeurs_tools = st.session_state.get("df_vendeurs", pd.DataFrame())
        df_agences_tools = st.session_state.get("df_agences", pd.DataFrame())
        df_directeurs_tools = st.session_state.get("df_directeurs", pd.DataFrame())
        target_sheet_evp = sheet_name_evp_for_next_month(periode_tools)
        st.caption(f"{periode_tools} → EVP {target_sheet_evp or 'à déterminer'}")

        evp_file = st.file_uploader(
            "Uploader EVP mensuelle.xlsx",
            type=["xlsx"],
            key=f"sidebar_evp_upload_{safe_filename(periode_tools)}"
        )

        if evp_file:
            try:
                evp_bytes, evp_error, evp_absents = update_evp_workbook(
                    evp_file,
                    df_vendeurs_tools,
                    df_directeurs_tools,
                    periode_tools
                )

                if evp_error:
                    st.error(f"❌ {evp_error}")
                else:
                    st.download_button(
                        "📥 Télécharger EVP",
                        evp_bytes,
                        f"EVP_mensuelle_mis_a_jour_{safe_filename(periode_tools)}.xlsx",
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key=f"sidebar_evp_download_{safe_filename(periode_tools)}",
                        on_click="ignore"
                    )

                    if evp_absents:
                        st.warning("Noms absents EVP : " + ", ".join(evp_absents))
                    else:
                        st.success("EVP prêt ✅")
            except Exception as exc:
                st.error(f"❌ EVP impossible : {exc}")

        st.divider()
        st.caption("📦 Exports PDF groupés")

        required_pdf_keys = [
            "col_client", "col_doc", "col_date", "col_op",
            "col_ca_magasin", "col_vente", "col_rem", "col_catalogue",
            "col_agence", "col_com1", "col_com2", "col_com3",
            "key_cols", "df_ok", "df_c"
        ]

        if all(k in st.session_state for k in required_pdf_keys):
            pdf_cols = {
                "client": st.session_state.col_client,
                "doc": st.session_state.col_doc,
                "date": st.session_state.col_date,
                "op": st.session_state.col_op,
                "ca_magasin": st.session_state.col_ca_magasin,
                "vente": st.session_state.col_vente,
                "rem": st.session_state.col_rem,
                "catalogue": st.session_state.col_catalogue,
                "agence": st.session_state.col_agence,
                "commerciaux": [
                    st.session_state.col_com1,
                    st.session_state.col_com2,
                    st.session_state.col_com3
                ],
            }

            bulk_pdf_items = []

            with st.spinner("Préparation PDF..."):
                for _, vendeur_row in df_vendeurs_tools.sort_values("Commercial").iterrows():
                    vendeur_nom = vendeur_row.get("Commercial", "")
                    detail_pdf, metrics_pdf = build_detail_vendeur_pdf(
                        vendeur_nom,
                        vendeur_row,
                        st.session_state.df_ok.copy(),
                        st.session_state.df_c.copy(),
                        pdf_cols["commerciaux"],
                        st.session_state.key_cols,
                        pdf_cols,
                        periode_tools
                    )

                    if detail_pdf is None:
                        continue

                    pdf_bytes = make_simple_pdf(
                        f"Detail des affaires - {vendeur_nom} - {periode_tools}",
                        metrics_pdf,
                        detail_pdf.reset_index(drop=True)
                    )
                    bulk_pdf_items.append((
                        f"vendeurs/{safe_filename(vendeur_nom)}_{safe_filename(periode_tools)}.pdf",
                        pdf_bytes
                    ))

                if not df_agences_tools.empty:
                    for _, agence_row in df_agences_tools.sort_values("agence").iterrows():
                        agence_nom = agence_row.get("agence", "")
                        detail_pdf, metrics_pdf = build_detail_agence_pdf(
                            agence_nom,
                            agence_row,
                            st.session_state.df_ok.copy(),
                            st.session_state.df_c.copy(),
                            st.session_state.key_cols,
                            pdf_cols,
                            periode_tools
                        )

                        if detail_pdf is None:
                            continue

                        pdf_bytes = make_simple_pdf(
                            f"Detail des affaires agence - {agence_nom} - {periode_tools}",
                            metrics_pdf,
                            detail_pdf.reset_index(drop=True)
                        )
                        bulk_pdf_items.append((
                            f"agences/{safe_filename(agence_nom)}_{safe_filename(periode_tools)}.pdf",
                            pdf_bytes
                        ))

            if bulk_pdf_items:
                zip_bytes = make_pdf_zip(bulk_pdf_items)
                st.download_button(
                    f"📦 Télécharger PDF ({len(bulk_pdf_items)})",
                    zip_bytes,
                    f"exports_pdf_{safe_filename(periode_tools)}.zip",
                    "application/zip",
                    key=f"sidebar_zip_pdf_all_{safe_filename(periode_tools)}",
                    on_click="ignore"
                )
            else:
                st.info("Aucun PDF à générer.")
        else:
            st.caption("Charge une période pour activer les exports.")


# ====================== BACK OFFICE USERS ======================

def afficher_admin_users():
    st.subheader("⚙️ Gestion des utilisateurs")

    users = load_users()

    st.write("### 👥 Utilisateurs existants")

    if users:
        df_users = pd.DataFrame.from_dict(users, orient="index")
        df_users["Mot de passe sécurisé"] = df_users.apply(
            lambda row: "Oui" if str(row.get("password_hash", "")).startswith("pbkdf2_sha256$") else "À migrer",
            axis=1
        )
        df_users["2FA"] = df_users.apply(
            lambda row: "Activée" if row.get("totp_enabled") and row.get("totp_confirmed") else ("À configurer" if row.get("totp_enabled") else "Non"),
            axis=1
        )
        df_users = df_users.drop(columns=["password", "password_hash", "totp_secret"], errors="ignore")
        df_users.index.name = "Identifiant"
        st.dataframe(df_users, use_container_width=True)
    else:
        st.info("Aucun utilisateur.")

    st.divider()

    st.write("### ➕ Ajouter un utilisateur")

    col1, col2 = st.columns(2)

    with col1:
        new_user = st.text_input("Nouvel identifiant", key="new_user").lower().strip()
        new_password = st.text_input("Mot de passe", type="password", key="new_password")

    with col2:
        new_role = st.selectbox("Rôle", ["admin", "vendeur", "directeur_agence"], key="new_role")
        new_nom = st.text_input("Nom vendeur / utilisateur", key="new_nom").upper().strip()
        new_agence = st.text_input("Agence si directeur", key="new_agence").upper().strip()
        new_totp_enabled = st.checkbox("Activer Authenticator (2FA)", key="new_totp_enabled")

    if st.button("Créer utilisateur", type="primary"):
        if not new_user or not new_password or not new_nom:
            st.error("Identifiant, mot de passe et nom sont obligatoires.")
        elif not password_is_strong_enough(new_password):
            st.error("Le mot de passe doit contenir au moins 8 caractères avec au moins une lettre et un chiffre.")
        elif new_user in users:
            st.error("Cet utilisateur existe déjà.")
        else:
            users[new_user] = {
                "password_hash": hash_password(new_password),
                "role": new_role,
                "nom": new_nom,
                "agence": new_agence if new_role == "directeur_agence" else None,
                "totp_enabled": bool(new_totp_enabled),
                "totp_secret": generate_totp_secret() if new_totp_enabled else "",
                "totp_confirmed": False
            }
            save_users(users)
            st.success("Utilisateur créé ✅")
            st.rerun()

    st.divider()

    st.write("### ✏️ Modifier un utilisateur")

    if users:
        user_edit = st.selectbox("Utilisateur à modifier", list(users.keys()), key="user_edit")

        if user_edit:
            u = users[user_edit]

            edit_password = st.text_input(
                "Nouveau mot de passe",
                value="",
                type="password",
                key="edit_password",
                help="Laisse vide pour conserver le mot de passe actuel."
            )

            role_options = ["admin", "vendeur", "directeur_agence"]
            current_role = u.get("role", "vendeur")
            role_index = role_options.index(current_role) if current_role in role_options else 1

            edit_role = st.selectbox("Rôle", role_options, index=role_index, key="edit_role")
            edit_nom = st.text_input("Nom vendeur / utilisateur", value=u.get("nom", ""), key="edit_nom").upper().strip()
            edit_agence = st.text_input("Agence", value=u.get("agence") or "", key="edit_agence").upper().strip()
            edit_totp_enabled = st.checkbox(
                "Activer Authenticator (2FA)",
                value=bool(u.get("totp_enabled")),
                key="edit_totp_enabled"
            )

            if edit_totp_enabled:
                current_secret = u.get("totp_secret") or generate_totp_secret()
                setup_uri = totp_uri(user_edit, current_secret)
                st.caption("QR code à scanner dans Google Authenticator ou Microsoft Authenticator.")
                st.image(qr_code_url(setup_uri), width=220)
                st.code(current_secret, language=None)

                if st.button("🔄 Régénérer le secret 2FA", key="regen_totp_secret"):
                    users[user_edit]["totp_secret"] = generate_totp_secret()
                    users[user_edit]["totp_enabled"] = True
                    users[user_edit]["totp_confirmed"] = False
                    save_users(users)
                    st.success("Secret 2FA régénéré ✅")
                    st.rerun()

            if st.button("Enregistrer modification"):
                updated_user = {
                    "role": edit_role,
                    "nom": edit_nom,
                    "agence": edit_agence if edit_role == "directeur_agence" else None,
                    "totp_enabled": bool(edit_totp_enabled),
                    "totp_secret": (u.get("totp_secret") or generate_totp_secret()) if edit_totp_enabled else "",
                    "totp_confirmed": bool(u.get("totp_confirmed")) if edit_totp_enabled else False
                }

                if edit_password:
                    if not password_is_strong_enough(edit_password):
                        st.error("Le mot de passe doit contenir au moins 8 caractères avec au moins une lettre et un chiffre.")
                        st.stop()
                    updated_user["password_hash"] = hash_password(edit_password)
                else:
                    if u.get("password_hash"):
                        updated_user["password_hash"] = u.get("password_hash")
                    elif u.get("password"):
                        updated_user["password_hash"] = hash_password(u.get("password"))

                users[user_edit] = updated_user
                save_users(users)

                if user_edit == st.session_state.get("username"):
                    st.session_state.user = users[user_edit]

                st.success("Utilisateur modifié ✅")
                st.rerun()

    st.divider()

    st.write("### ❌ Supprimer un utilisateur")

    users = load_users()

    if users:
        user_delete = st.selectbox("Utilisateur à supprimer", list(users.keys()), key="user_delete")

        if st.button("Supprimer utilisateur"):
            if user_delete == st.session_state.get("username"):
                st.error("Tu ne peux pas supprimer ton propre compte connecté.")
            elif user_delete == "joseph":
                st.error("Impossible de supprimer l’admin principal Joseph.")
            else:
                del users[user_delete]
                save_users(users)
                st.success("Utilisateur supprimé ✅")
                st.rerun()


# ====================== AFFICHAGE ANNUEL ======================

def afficher_annuel(tab):
    with tab:
        st.subheader("📆 Analyse annuelle")

        df_all_vendeurs, df_all_agences, df_all_directeurs = load_all_historique()

        if df_all_vendeurs.empty and df_all_agences.empty:
            st.info("Aucune période historique disponible pour construire l’analyse annuelle.")
            return

        annees_detectees = []
        for df_src in [df_all_vendeurs, df_all_agences]:
            if not df_src.empty and "annee" in df_src.columns:
                annees_detectees += [int(a) for a in df_src["annee"].dropna().unique()]

        annees = sorted(set(annees_detectees))

        if not annees:
            st.warning("Impossible de détecter les années depuis les périodes sauvegardées. Utilise un format du type : Avril 2026.")
            return

        default_year = datetime.today().year if datetime.today().year in annees else max(annees)

        annee_selectionnee = st.selectbox(
            "Année analysée",
            annees,
            index=annees.index(default_year)
        )

        periodes_comptables = [
            p for p in sorted(list_periodes(), key=periode_sort_key)
            if is_periode_comptable_annuelle(p, annee_selectionnee, use_m2_rule)
        ]

        if periodes_comptables:
            if use_m2_rule:
                st.caption("Mois comptabilisés selon la règle M-2 : " + ", ".join(periodes_comptables))
            else:
                st.warning("⚠️ Mode TEST actif : toutes les périodes de l’année sélectionnée sont comptabilisées.")
                st.caption("Mois comptabilisés : " + ", ".join(periodes_comptables))
        else:
            st.warning("Aucun mois comptabilisable pour cette année.")
            return

        df_v = df_all_vendeurs[
            df_all_vendeurs["periode"].isin(periodes_comptables)
        ].copy() if not df_all_vendeurs.empty else pd.DataFrame()

        df_a = df_all_agences[
            df_all_agences["periode"].isin(periodes_comptables)
        ].copy() if not df_all_agences.empty else pd.DataFrame()

        df_v_classement_all = df_v.copy()

        if role == "vendeur":
            df_v = df_v[
                df_v["Commercial"].apply(normalize_key) == normalize_key(user["nom"])
            ].copy() if not df_v.empty else pd.DataFrame()
            df_a = pd.DataFrame()

        elif role == "directeur_agence":
            df_v = df_v[
                df_v["Commercial"].apply(normalize_key) == normalize_key(user["nom"])
            ].copy() if not df_v.empty else pd.DataFrame()

            if not df_a.empty:
                df_a = df_a[
                    df_a["agence"].apply(normalize_key) == normalize_key(user["agence"])
                ].copy()

        total_ok_annuel = df_v["ca_ok"].sum() if not df_v.empty and "ca_ok" in df_v.columns else 0
        total_commissions_annuel = df_v["commission_eur"].sum() if not df_v.empty and "commission_eur" in df_v.columns else 0
        total_agence_ok_annuel = df_a["ca_ok"].sum() if not df_a.empty and "ca_ok" in df_a.columns else 0

        c1, c2, c3 = st.columns(3)

        with c1:
            card("✅ CA OK vendeurs annuel", f"{total_ok_annuel:,.2f} €")

        with c2:
            card("💰 Commissions vendeurs", f"{total_commissions_annuel:,.2f} €")

        with c3:
            card("🏢 CA OK agences annuel", f"{total_agence_ok_annuel:,.2f} €")

        st.divider()

        if not df_v_classement_all.empty:
            df_classement_source = df_v_classement_all[
                ~df_v_classement_all["Commercial"].apply(is_responsable_agence)
            ].copy()
            df_classement_vendeurs = (
                df_classement_source
                .groupby("Commercial", as_index=False)
                .agg({
                    "ca_ok": "sum",
                    "commission_eur": "sum"
                })
                .sort_values("ca_ok", ascending=False)
                .reset_index(drop=True)
            )

            df_classement_vendeurs.insert(0, "Rang", range(1, len(df_classement_vendeurs) + 1))

            st.subheader("🏆 Classement annuel vendeurs — CA OK")
            st.dataframe(
                df_classement_vendeurs.rename(columns={
                    "ca_ok": "CA OK annuel",
                    "commission_eur": "Commission annuelle"
                }),
                use_container_width=True
            )

            df_mensuel_vendeurs = (
                df_v
                .groupby(["periode", "mois_num"], as_index=False)
                .agg({
                    "ca_ok": "sum",
                    "commission_eur": "sum"
                })
                .sort_values("mois_num")
            )

            st.subheader("📈 Évolution mensuelle vendeurs")
            st.dataframe(
                df_mensuel_vendeurs.rename(columns={
                    "periode": "Période",
                    "ca_ok": "CA OK",
                    "commission_eur": "Commission"
                }).drop(columns=["mois_num"], errors="ignore"),
                use_container_width=True
            )
        else:
            st.info("Aucune donnée vendeur pour cette analyse annuelle.")

        if role in ["admin", "directeur_agence"]:
            st.divider()

            if not df_a.empty:
                df_classement_agences = (
                    df_a
                    .groupby("agence", as_index=False)
                    .agg({
                        "ca_ok": "sum",
                        "ca_magasin_ok": "sum",
                        "nb_ok": "sum",
                        "nb_total": "sum"
                    })
                    .sort_values("ca_ok", ascending=False)
                    .reset_index(drop=True)
                )

                df_classement_agences.insert(0, "Rang", range(1, len(df_classement_agences) + 1))

                st.subheader("🏢 Classement annuel agences — CA OK")
                st.dataframe(
                    df_classement_agences.rename(columns={
                        "agence": "Agence",
                        "ca_ok": "CA OK annuel",
                        "ca_magasin_ok": "CA magasin OK annuel",
                        "nb_ok": "Nb affaires OK",
                        "nb_total": "Nb affaires total"
                    }),
                    use_container_width=True
                )

                df_mensuel_agences = (
                    df_a
                    .groupby(["periode", "mois_num"], as_index=False)
                    .agg({
                        "ca_ok": "sum",
                        "ca_magasin_ok": "sum",
                        "nb_ok": "sum"
                    })
                    .sort_values("mois_num")
                )

                st.subheader("📈 Évolution mensuelle agences")
                st.dataframe(
                    df_mensuel_agences.rename(columns={
                        "periode": "Période",
                        "ca_ok": "CA OK",
                        "ca_magasin_ok": "CA magasin OK",
                        "nb_ok": "Nb affaires OK"
                    }).drop(columns=["mois_num"], errors="ignore"),
                    use_container_width=True
                )
            else:
                st.info("Aucune donnée agence pour cette analyse annuelle.")


def afficher_dossiers_en_attente(tab):
    with tab:
        st.subheader("⏳ Dossiers en attente")

        df_ok = st.session_state.df_ok.copy()
        df_c = st.session_state.df_c.copy()

        col_client = st.session_state.col_client
        col_doc = st.session_state.col_doc
        col_date = st.session_state.col_date
        col_op = st.session_state.col_op
        col_agence = st.session_state.col_agence
        col_vente = st.session_state.col_vente
        col_ca_magasin = st.session_state.col_ca_magasin
        col_catalogue = st.session_state.col_catalogue
        col_rem = st.session_state.col_rem
        colonnes_commerciaux = [
            st.session_state.col_com1,
            st.session_state.col_com2,
            st.session_state.col_com3
        ]

        attente = remove_attente_already_ok(
            df_ok,
            df_c,
            st.session_state.key_cols,
            col_client,
            col_agence,
            col_vente,
            col_ca_magasin,
            col_catalogue
        )

        if attente.empty:
            st.success("Aucun dossier en attente.")
            return

        attente = attente.copy()
        attente["Commerciaux"] = attente.apply(
            lambda row: list_commerciaux_row(row, colonnes_commerciaux),
            axis=1
        )
        attente["_ATTENTE_KEY_"] = attente.apply(
            lambda row: make_attente_tracking_key(row, col_client, col_doc, col_agence, col_ca_magasin, col_catalogue),
            axis=1
        )

        motifs_attente = load_motifs_attente()
        attente["Motif d'attente"] = attente["_ATTENTE_KEY_"].apply(
            lambda key: clean_visible(motifs_attente.get(key, {}).get("motif", ""))
        )
        attente["Détail motif"] = attente["_ATTENTE_KEY_"].apply(
            lambda key: clean_visible(motifs_attente.get(key, {}).get("detail", ""))
        )
        attente["Dernière relance"] = attente["_ATTENTE_KEY_"].apply(
            lambda key: clean_visible(motifs_attente.get(key, {}).get("derniere_relance", ""))
        )

        if col_op and col_op in attente.columns:
            attente["OPC"] = np.where(attente.apply(lambda row: is_opc(row, col_op), axis=1), "OUI", "")
        else:
            attente["OPC"] = ""

        if col_catalogue and col_catalogue in attente.columns and col_rem and col_rem in attente.columns:
            catalogue = pd.to_numeric(attente[col_catalogue], errors="coerce").fillna(0)
            remise = pd.to_numeric(attente[col_rem], errors="coerce").fillna(0)
            attente["Remise %"] = np.where(catalogue > 0, remise / catalogue * 100, 0)
        else:
            attente["Remise %"] = 0

        vendeurs = sorted({
            nom
            for value in attente["Commerciaux"].dropna()
            for nom in [n.strip() for n in str(value).split("/")]
            if nom
        })
        agences = sorted(attente[col_agence].dropna().map(clean_visible).unique()) if col_agence in attente.columns else []

        f1, f2, f3, f4, f5 = st.columns([2, 2, 1, 2, 2])

        vendeur_filtre = f1.selectbox("Commercial", ["Tous"] + vendeurs, key="attente_filtre_vendeur")
        agence_filtre = f2.selectbox("Agence", ["Toutes"] + agences, key="attente_filtre_agence")
        opc_filtre = f3.selectbox("OPC", ["Tous", "Oui", "Non"], key="attente_filtre_opc")
        recherche = f4.text_input("Recherche client / document", key="attente_recherche").strip()
        motif_filtre = f5.selectbox(
            "Motif",
            ["Tous", "À compléter"] + [m for m in MOTIFS_ATTENTE_OPTIONS if m],
            key="attente_filtre_motif"
        )

        filtered = attente.copy()

        if vendeur_filtre != "Tous":
            filtered = filtered[
                filtered["Commerciaux"].apply(lambda value: normalize_key(vendeur_filtre) in [normalize_key(n) for n in str(value).split("/")])
            ]

        if agence_filtre != "Toutes" and col_agence in filtered.columns:
            filtered = filtered[filtered[col_agence].apply(normalize_key) == normalize_key(agence_filtre)]

        if opc_filtre == "Oui":
            filtered = filtered[filtered["OPC"] == "OUI"]
        elif opc_filtre == "Non":
            filtered = filtered[filtered["OPC"] != "OUI"]

        if motif_filtre == "À compléter":
            filtered = filtered[filtered["Motif d'attente"].eq("")]
        elif motif_filtre != "Tous":
            filtered = filtered[filtered["Motif d'attente"].apply(normalize_key) == normalize_key(motif_filtre)]

        if recherche:
            search_cols = [c for c in [col_client, col_doc, "Commerciaux", col_agence] if c and c in filtered.columns]
            search_key = normalize_key(recherche)
            mask = pd.Series(False, index=filtered.index)
            for c in search_cols:
                mask |= filtered[c].apply(lambda value: search_key in normalize_key(value))
            filtered = filtered[mask]

        sort_cols = [c for c in [col_agence, "Commerciaux", col_client] if c and c in filtered.columns]
        if sort_cols:
            filtered = (
                filtered
                .assign(**{f"_sort_{c}": filtered[c].apply(normalize_key) for c in sort_cols})
                .sort_values([f"_sort_{c}" for c in sort_cols])
                .drop(columns=[f"_sort_{c}" for c in sort_cols])
            )

        total_ca_global = (
            pd.to_numeric(filtered[col_ca_magasin], errors="coerce").fillna(0).sum()
            if col_ca_magasin in filtered.columns
            else 0
        )

        dossiers_sans_motif = int(filtered["Motif d'attente"].eq("").sum()) if "Motif d'attente" in filtered.columns else 0

        c1, c2, c3 = st.columns(3)
        with c1:
            card("⏳ Nb dossiers", f"{len(filtered)}")
        with c2:
            card("🏢 CA global en attente", f"{total_ca_global:,.2f} €")
        with c3:
            card("📝 Motif à compléter", f"{dossiers_sans_motif}")

        if dossiers_sans_motif:
            st.caption("Motifs manquants à compléter depuis le volet de suivi ci-dessous.")

        filtered_reset = filtered.reset_index(drop=True)
        selected_rows = []
        table_state = st.session_state.get("attente_table_selection")
        try:
            selected_rows = list(table_state.selection.rows)
        except Exception:
            if isinstance(table_state, dict):
                selected_rows = list(table_state.get("selection", {}).get("rows", []))

        selected_key_from_table = None
        if selected_rows:
            selected_pos = selected_rows[0]
            if 0 <= selected_pos < len(filtered_reset):
                selected_key_from_table = filtered_reset.iloc[selected_pos].get("_ATTENTE_KEY_", "")

        if not filtered.empty:
            dossier_options = []
            dossier_labels = {}

            for display_idx, row in filtered_reset.iterrows():
                key = row.get("_ATTENTE_KEY_", "")
                client_label = clean_visible(row.get(col_client, "")) if col_client else "Dossier"
                doc_label = clean_visible(row.get(col_doc, "")) if col_doc else ""
                agence_label = clean_visible(row.get(col_agence, "")) if col_agence else ""
                motif_label = clean_visible(row.get("Motif d'attente", "")) or "À compléter"
                label = f"{client_label}"
                if doc_label:
                    label += f" | {doc_label}"
                if agence_label:
                    label += f" | {agence_label}"
                label += f" | {motif_label}"
                option_key = f"{key}__{display_idx}"
                dossier_options.append(option_key)
                dossier_labels[option_key] = label

            selected_option_from_table = None
            if selected_key_from_table:
                for option in dossier_options:
                    if option.rsplit("__", 1)[0] == selected_key_from_table:
                        selected_option_from_table = option
                        break

            select_key = "attente_dossier_motif_select"
            if selected_option_from_table:
                st.session_state[select_key] = selected_option_from_table
            elif st.session_state.get(select_key) not in dossier_options:
                st.session_state[select_key] = dossier_options[0]

            st.markdown('<div id="motif-editor-anchor"></div>', unsafe_allow_html=True)
            if selected_key_from_table:
                components.html(
                    """
                    <script>
                    const scrollToEditor = () => {
                        try {
                            const anchor = window.parent.document.getElementById("motif-editor-anchor");
                            if (anchor) {
                                anchor.scrollIntoView({ behavior: "smooth", block: "start" });
                            }
                        } catch (error) {}
                    };
                    setTimeout(scrollToEditor, 120);
                    </script>
                    """,
                    height=0,
                )

            with st.expander(
                "📝 Modifier un motif d'attente / relance",
                expanded=bool(selected_key_from_table)
            ):
                selected_option = st.selectbox(
                    "Dossier à mettre à jour",
                    dossier_options,
                    format_func=lambda key: dossier_labels.get(key, key),
                    key=select_key
                )
                selected_key = selected_option.rsplit("__", 1)[0]
                selected_row = filtered[filtered["_ATTENTE_KEY_"] == selected_key].iloc[0]
                selected_record = motifs_attente.get(selected_key, {})
                selected_suffix = safe_filename(selected_key)[0:80]

                motif_current = clean_visible(selected_record.get("motif", ""))
                motif_index = MOTIFS_ATTENTE_OPTIONS.index(motif_current) if motif_current in MOTIFS_ATTENTE_OPTIONS else 0

                with st.form(key=f"form_motif_attente_{selected_suffix}"):
                    m1, m2 = st.columns([1, 2])
                    with m1:
                        motif_value = st.selectbox(
                            "Motif d'attente obligatoire",
                            MOTIFS_ATTENTE_OPTIONS,
                            index=motif_index
                        )
                        relance_value = st.checkbox(
                            "Relance effectuée aujourd'hui",
                            value=False
                        )
                    with m2:
                        detail_value = st.text_area(
                            "Détail / élément manquant",
                            value=clean_visible(selected_record.get("detail", "")),
                            height=80
                        )

                    save_motif = st.form_submit_button("💾 Enregistrer le suivi")

                if save_motif:
                    if not clean_visible(motif_value):
                        st.error("Le motif d'attente est obligatoire.")
                    else:
                        previous_relance = clean_visible(selected_record.get("derniere_relance", ""))
                        motifs_attente[selected_key] = {
                            "motif": clean_visible(motif_value),
                            "detail": clean_visible(detail_value),
                            "derniere_relance": datetime.now().strftime("%d/%m/%Y") if relance_value else previous_relance,
                            "updated_at": datetime.now().isoformat(timespec="seconds"),
                            "updated_by": user.get("nom", st.session_state.get("username", ""))
                        }
                        save_motifs_attente(motifs_attente)
                        st.success("Motif d'attente enregistré ✅")
                        st.rerun()

                mail_subject = f"Éléments manquants dossier {clean_visible(selected_row.get(col_client, ''))}"
                mail_body = (
                    "Bonjour,\n\n"
                    "Nous revenons vers vous concernant votre dossier.\n\n"
                    f"Motif d'attente : {clean_visible(motif_current) or '[à compléter]'}\n"
                    f"Élément attendu : {clean_visible(selected_record.get('detail', '')) or '[à compléter]'}\n\n"
                    "Merci de nous transmettre les éléments nécessaires afin que nous puissions faire avancer le dossier.\n\n"
                    "Cordialement,\nEcoHabitat"
                )
                mailto_url = f"mailto:?subject={quote(mail_subject)}&body={quote(mail_body)}"
                st.link_button("📧 Préparer un mail de relance", mailto_url)

        cols_show = [
            col_client,
            "Commerciaux",
            col_doc,
            col_date,
            col_agence,
            col_ca_magasin,
            "Remise %",
            "OPC",
            "Motif d'attente",
            "Détail motif"
        ]
        cols_show = list(dict.fromkeys([c for c in cols_show if c and c in filtered.columns]))
        affichage = filtered[cols_show].copy()

        rename_cols = {}
        if col_client:
            rename_cols[col_client] = "Client / Référence affaire"
        if col_doc:
            rename_cols[col_doc] = "N° Document"
        if col_date:
            rename_cols[col_date] = "Date document"
        if col_agence:
            rename_cols[col_agence] = "Agence"
        if col_vente:
            rename_cols[col_vente] = "TOTAL VENTE"
        if col_ca_magasin:
            rename_cols[col_ca_magasin] = "CA global affaire"
        affichage = affichage.rename(columns=rename_cols)

        for c in ["TOTAL VENTE", "CA global affaire", "Remise %"]:
            if c in affichage.columns:
                affichage[c] = pd.to_numeric(affichage[c], errors="coerce").fillna(0).round(2)

        if "Date document" in affichage.columns:
            affichage["Date document"] = pd.to_datetime(
                affichage["Date document"],
                errors="coerce"
            ).dt.strftime("%d/%m/%Y").fillna("")

        st.dataframe(
            affichage.reset_index(drop=True),
            use_container_width=True,
            height=650,
            key="attente_table_selection",
            on_select="rerun",
            selection_mode="single-row"
        )

        if not filtered.empty:
            with st.expander("📧 Mail interne hebdomadaire", expanded=False):
                settings = load_settings()
                internal_recipients = st.text_input(
                    "Destinataires internes",
                    value=clean_visible(settings.get("attente_internal_recipients", "")),
                    placeholder="secretariat@...; associe@...",
                    key="attente_internal_recipients"
                )

                if st.button("💾 Enregistrer les destinataires", key="save_attente_internal_recipients"):
                    settings["attente_internal_recipients"] = clean_visible(internal_recipients)
                    save_settings(settings)
                    st.success("Destinataires enregistrés ✅")

                internal_subject = f"Suivi hebdomadaire dossiers en attente - {periode}"
                internal_body = build_attente_internal_mail_body(
                    filtered,
                    periode,
                    col_client,
                    col_doc,
                    col_date,
                    col_agence,
                    col_ca_magasin
                )
                recipients_url = quote(clean_visible(internal_recipients), safe="@.;,")
                internal_mailto = f"mailto:{recipients_url}?subject={quote(internal_subject)}&body={quote(internal_body)}"

                st.link_button("📨 Préparer le mail interne", internal_mailto)
                st.download_button(
                    "📄 Télécharger le récap en TXT",
                    internal_body.encode("utf-8"),
                    f"suivi_attente_{safe_filename(periode)}.txt",
                    "text/plain",
                    key="download_attente_internal_summary"
                )


# ====================== AFFICHAGE DONNÉES ======================

if st.session_state.get("df_vendeurs") is not None:

    required_keys = [
        "df_ok", "df_c",
        "col_client", "col_doc", "col_date",
        "col_op", "col_ca_magasin", "col_vente",
        "col_rem", "col_catalogue",
        "col_agence", "col_com1", "col_com2", "col_com3",
        "key_cols"
    ]

    missing_keys = [k for k in required_keys if k not in st.session_state]

    if missing_keys:
        st.warning(
            "⚠️ Cette période a été sauvegardée avec une ancienne version. "
            "Demande à un admin de recharger les fichiers CONFIRM et BONLIVR puis de relancer le traitement."
        )
        st.stop()

    periode = st.session_state.get("periode", "Mois inconnu")
    indicateurs_cache_key = (
        periode,
        id(st.session_state.df_vendeurs),
        id(st.session_state.get("df_agences", None)),
        id(st.session_state.df_ok),
        id(st.session_state.df_c),
        st.session_state.col_vente,
        st.session_state.col_catalogue,
        st.session_state.col_rem,
        st.session_state.col_op,
        st.session_state.col_client,
        st.session_state.col_agence,
        st.session_state.col_ca_magasin,
        st.session_state.col_com1,
        st.session_state.col_com2,
        st.session_state.col_com3,
        tuple(st.session_state.key_cols),
    )
    indicateurs_cache = st.session_state.get("_indicateurs_cache", {})

    if indicateurs_cache.get("key") == indicateurs_cache_key:
        df_vendeurs_all = indicateurs_cache["df_vendeurs"].copy()
        df_agences_all = indicateurs_cache["df_agences"].copy()
    else:
        df_vendeurs_all = recompute_df_vendeurs_indicators(
            st.session_state.df_vendeurs.copy(),
            st.session_state.df_ok.copy(),
            st.session_state.df_c.copy(),
            st.session_state.col_vente,
            st.session_state.col_catalogue,
            st.session_state.col_rem,
            st.session_state.col_op,
            [
                st.session_state.col_com1,
                st.session_state.col_com2,
                st.session_state.col_com3
            ],
            st.session_state.key_cols,
            st.session_state.col_client,
            st.session_state.col_agence,
            st.session_state.col_ca_magasin
        )
        df_agences_all = st.session_state.get("df_agences", pd.DataFrame()).copy()
        df_agences_all = recompute_df_agences_attente(
            df_agences_all,
            st.session_state.df_ok.copy(),
            st.session_state.df_c.copy(),
            st.session_state.key_cols,
            st.session_state.col_client,
            st.session_state.col_agence,
            st.session_state.col_vente,
            st.session_state.col_ca_magasin,
            st.session_state.col_catalogue,
            st.session_state.col_op,
            [
                st.session_state.col_com1,
                st.session_state.col_com2,
                st.session_state.col_com3
            ]
        )
        st.session_state["_indicateurs_cache"] = {
            "key": indicateurs_cache_key,
            "df_vendeurs": df_vendeurs_all.copy(),
            "df_agences": df_agences_all.copy(),
        }

    df_directeurs_all = st.session_state.get("df_directeurs", pd.DataFrame()).copy()

    # ====================== FILTRAGE ROLE ======================

    if role == "vendeur":
        df_vendeurs = df_vendeurs_all[
            df_vendeurs_all["Commercial"].apply(normalize_key) == normalize_key(user["nom"])
        ].copy()

        df_agences = pd.DataFrame()
        df_directeurs = pd.DataFrame()

    elif role == "directeur_agence":
        df_vendeurs = df_vendeurs_all[
            df_vendeurs_all["Commercial"].apply(normalize_key) == normalize_key(user["nom"])
        ].copy()

        df_agences = df_agences_all[
            df_agences_all["agence"].apply(normalize_key) == normalize_key(user["agence"])
        ].copy() if not df_agences_all.empty else pd.DataFrame()

        df_directeurs = df_directeurs_all[
            df_directeurs_all["directeur"].apply(normalize_key) == normalize_key(user["nom"])
        ].copy() if not df_directeurs_all.empty else pd.DataFrame()

    else:
        df_vendeurs = df_vendeurs_all
        df_agences = df_agences_all
        df_directeurs = df_directeurs_all

    periode_title_col, periode_select_col = st.columns([2, 1])
    with periode_title_col:
        st.subheader(f"📅 Période : **{periode}**")

    periodes_main = sorted(list_periodes(), key=periode_sort_key)
    if periodes_main:
        current_periode_index = periodes_main.index(periode) if periode in periodes_main else 0
        with periode_select_col:
            periode_main_select = st.selectbox(
                "Changer de période",
                periodes_main,
                index=current_periode_index,
                key="periode_main_select"
            )

        if periode_main_select != periode:
            if load_periode_preserve_ui(periode_main_select):
                st.rerun()
            else:
                st.error("❌ Impossible de charger cette période.")

    if role == "admin":
        pages = [
            "📊 Dashboard",
            "⏳ En attente",
            "📆 Annuel",
            "👤 Par Vendeur",
            "🏢 Par Agence",
            "👔 Directeurs",
            "📋 Listes complètes",
            "⚙️ Utilisateurs"
        ]
        active_page = st.pills(
            "Navigation",
            pages,
            default=pages[0],
            label_visibility="collapsed",
            key="active_page_admin",
            width="stretch"
        )
        active_page = active_page or st.session_state.get("active_page_admin") or pages[0]

    elif role == "directeur_agence":
        pages = [
            "👤 Mes chiffres",
            "📆 Annuel",
            "🏢 Mon agence",
            "👔 Commission agence"
        ]
        active_page = st.pills(
            "Navigation",
            pages,
            default=pages[0],
            label_visibility="collapsed",
            key="active_page_directeur",
            width="stretch"
        )
        active_page = active_page or st.session_state.get("active_page_directeur") or pages[0]

    else:
        pages = [
            "👤 Mes chiffres",
            "📆 Annuel"
        ]
        active_page = st.pills(
            "Navigation",
            pages,
            default=pages[0],
            label_visibility="collapsed",
            key="active_page_vendeur",
            width="stretch"
        )
        active_page = active_page or st.session_state.get("active_page_vendeur") or pages[0]

    # ====================== FONCTIONS AFFICHAGE ======================

    def afficher_vendeur(tab, vendeur_forced=None):
        with tab:
            if df_vendeurs.empty:
                st.info("Aucune donnée vendeur disponible pour ce compte.")
                return

            if vendeur_forced:
                vendeur = vendeur_forced
            else:
                vendeurs_options = sorted(df_vendeurs["Commercial"])
                vendeur_precedent = st.session_state.get("vendeur_select")
                vendeur_index = vendeurs_options.index(vendeur_precedent) if vendeur_precedent in vendeurs_options else 0
                vendeur = st.selectbox(
                    "Sélectionner un commercial",
                    vendeurs_options,
                    index=vendeur_index,
                    key="vendeur_select"
                )

            data = df_vendeurs[df_vendeurs["Commercial"].apply(normalize_key) == normalize_key(vendeur)]

            if data.empty:
                st.info("Aucune donnée trouvée pour ce vendeur.")
                return

            data = data.iloc[0]

            projection_total = st.toggle(
                "Afficher la commission sur le CA Total",
                value=False,
                key=f"projection_total_{safe_filename(vendeur)}"
            )

            if projection_total:
                if is_responsable_agence(vendeur):
                    base_affichee = 10
                    points_affiches = 0
                    commission_pct_affichee = 10
                    ca_projection_directeur = to_float(
                        data.get(
                            "ca_commissionnable_directeur_total",
                            data.get("ca_total", 0)
                        )
                    )
                    commission_eur_affichee = round(ca_projection_directeur * 0.10, 2)
                else:
                    remise_commission_affichee = to_float(data.get("remise_hors_opc_global_pct", data.get("remise_global_pct", 0)))
                    base_affichee, points_affiches, commission_pct_affichee, commission_eur_affichee = calculate_commission(
                        to_float(data.get("ca_total", 0)),
                        remise_commission_affichee
                    )
                commission_card_label = "💰 Commission projetée"
                mode_commission_label = "Projection sur CA total"
            else:
                remise_commission_affichee = to_float(data.get("remise_hors_opc_pct", 0))
                base_affichee = to_float(data.get("base_commission_pct", 0))
                points_affiches = to_float(data.get("points_perdus", 0))
                commission_pct_affichee = to_float(data.get("commission_pct", 0))
                commission_eur_affichee = to_float(data.get("commission_eur", 0))
                commission_card_label = "💰 Commission"
                mode_commission_label = "Rémunération réelle sur CA OK"

            c1, c2, c3, c4, c5 = st.columns(5)

            with c1:
                card("✅ CA OK", f"{data['ca_ok']:,.2f} €")

            with c2:
                card("⏳ CA en attente", f"{data['ca_attente']:,.2f} €")

            with c3:
                card("📌 CA Total", f"{data['ca_total']:,.2f} €")

            with c4:
                card(commission_card_label, f"{commission_eur_affichee:,.2f} €")

            with c5:
                ratio_opc_vendeur = to_float(data.get("ratio_opc_total_pct", 0))
                card("🧾 Ratio OPC", f"{ratio_opc_vendeur:,.1f} %")

            commission_summary_bar(
                base_affichee,
                points_affiches,
                commission_pct_affichee
            )

            if is_responsable_agence(vendeur):
                nb_remise_plus_30 = int(to_float(
                    data.get(
                        "nb_remise_plus_30_total" if projection_total else "nb_remise_plus_30_ok",
                        0
                    )
                ))
                if nb_remise_plus_30 > 0:
                    st.warning(
                        f"⚠️ {nb_remise_plus_30} dossier(s) avec une remise supérieure à 30 %. "
                        "La commission directeur reste calculée à 10 %, à contrôler."
                    )

            vendeur_bonus_malus_global = st.toggle(
                "Afficher la remise et le BONUS/MALUS sur le CA Total",
                value=False,
                key=f"vendeur_bonus_malus_global_{safe_filename(vendeur)}"
            )
            vendeur_bonus_malus_col = "bonus_malus_global" if vendeur_bonus_malus_global else "bonus_malus_ok"
            vendeur_bonus_malus_label = "🌍 Bonus / Malus Global" if vendeur_bonus_malus_global else "🎯 Bonus / Malus OK"
            vendeur_remise_col = "remise_global_pct" if vendeur_bonus_malus_global else "remise_ok_pct"
            vendeur_remise_label = "🌍 Remise Global hors OPC" if vendeur_bonus_malus_global else "🎯 Remise OK hors OPC"

            c_bonus1, c_bonus2 = st.columns(2)
            with c_bonus1:
                card(vendeur_remise_label, f"{to_float(data.get(vendeur_remise_col, 0)):,.2f} %")
            with c_bonus2:
                card(vendeur_bonus_malus_label, f"{to_float(data.get(vendeur_bonus_malus_col, 0)):,.2f} €")

            st.subheader(f"📋 Détail des affaires de **{vendeur}**")

            df_ok = st.session_state.df_ok.copy()
            df_c = st.session_state.df_c.copy()

            colonnes_commerciaux = [
                st.session_state.col_com1,
                st.session_state.col_com2,
                st.session_state.col_com3
            ]

            key_cols = st.session_state.key_cols
            col_client = st.session_state.col_client
            col_agence = st.session_state.col_agence
            col_vente = st.session_state.col_vente
            col_ca_magasin = st.session_state.col_ca_magasin
            col_catalogue = st.session_state.col_catalogue

            ok_detail = df_ok[vendeur_mask(df_ok, vendeur, colonnes_commerciaux)].copy()
            attente_detail = df_c[vendeur_mask(df_c, vendeur, colonnes_commerciaux)].copy()

            attente_detail = remove_attente_already_ok(
                ok_detail,
                attente_detail,
                key_cols,
                col_client,
                col_agence,
                col_vente,
                col_ca_magasin,
                col_catalogue
            )

            ok_detail["Statut"] = "✅ OK"
            attente_detail["Statut"] = "⏳ En attente"

            detail = pd.concat([ok_detail, attente_detail], ignore_index=True)

            if not detail.empty:
                col_doc = st.session_state.col_doc
                col_date = st.session_state.col_date
                col_rem = st.session_state.col_rem
                col_op = st.session_state.col_op
                motifs_attente = load_motifs_attente()

                afficher_alerte_hors_periode(detail, col_client, col_doc, col_date, periode)

                detail_calc = detail.copy()
                detail_calc["_ATTENTE_KEY_"] = detail_calc.apply(
                    lambda row: make_attente_tracking_key(row, col_client, col_doc, col_agence, col_ca_magasin, col_catalogue),
                    axis=1
                )
                detail_calc["Motif d'attente"] = detail_calc.apply(
                    lambda row: clean_visible(motifs_attente.get(row.get("_ATTENTE_KEY_", ""), {}).get("motif", ""))
                    if str(row.get("Statut", "")).startswith("⏳") else "",
                    axis=1
                )
                detail_calc["Détail motif"] = detail_calc.apply(
                    lambda row: clean_visible(motifs_attente.get(row.get("_ATTENTE_KEY_", ""), {}).get("detail", ""))
                    if str(row.get("Statut", "")).startswith("⏳") else "",
                    axis=1
                )
                detail_calc["_BULLE_ATTENTE_"] = detail_calc.apply(
                    lambda row: " | ".join([
                        part for part in [
                            clean_visible(row.get("Motif d'attente", "")),
                            clean_visible(row.get("Détail motif", ""))
                        ]
                        if part
                    ]),
                    axis=1
                )
                def statut_avec_motif(row):
                    motif = clean_visible(row.get("Motif d'attente", ""))
                    if motif:
                        return "⏳ En attente ⓘ"
                    return row.get("Statut", "")

                detail_calc["Statut"] = detail_calc.apply(statut_avec_motif, axis=1)

                for c in [col_vente, col_ca_magasin, col_catalogue, col_rem]:
                    if c and c in detail_calc.columns:
                        detail_calc[c] = pd.to_numeric(detail_calc[c], errors="coerce").fillna(0)

                detail_calc["Commerciaux"] = detail_calc.apply(
                    lambda row: list_commerciaux_row(row, colonnes_commerciaux),
                    axis=1
                )

                detail_calc["Nombre de vendeurs"] = detail_calc.apply(
                    lambda row: count_vendeurs_row(row, colonnes_commerciaux),
                    axis=1
                )

                if col_catalogue and col_catalogue in detail_calc.columns and col_rem and col_rem in detail_calc.columns:
                    detail_calc["Remise %"] = np.where(
                        detail_calc[col_catalogue] > 0,
                        detail_calc[col_rem] / detail_calc[col_catalogue] * 100,
                        0
                    )
                else:
                    detail_calc["Remise %"] = 0

                # Bonus / Malus par dossier :
                # base théorique = Total ventes avant remise avec 15 % de remise, réparti par nombre de vendeurs.
                # règle OPC : les remises OPC ne comptent pas dans la remise moyenne ;
                # pour le bonus/malus, on neutralise les malus OPC mais on conserve les bonus éventuels.
                if col_catalogue and col_catalogue in detail_calc.columns and col_vente and col_vente in detail_calc.columns:
                    objectif_15_par_vendeur = (detail_calc[col_catalogue] * 0.85) / detail_calc["Nombre de vendeurs"]
                    detail_calc["Bonus / Malus"] = detail_calc[col_vente] - objectif_15_par_vendeur
                else:
                    detail_calc["Bonus / Malus"] = 0

                if col_op and col_op in detail_calc.columns:
                    opc_mask = detail_calc.apply(lambda row: is_opc(row, col_op), axis=1)
                    detail_calc.loc[opc_mask & (detail_calc["Bonus / Malus"] < 0), "Bonus / Malus"] = 0
                    detail_calc["OPC"] = np.where(opc_mask, "OUI", "")
                else:
                    opc_mask = pd.Series(False, index=detail_calc.index)
                    detail_calc["OPC"] = ""

                cols_show = [
                    col_client,
                    "Commerciaux",
                    col_doc,
                    col_date,
                    "Statut",
                    col_agence,
                    col_vente,
                    col_ca_magasin,
                    "Remise %",
                    "OPC",
                    "Bonus / Malus"
                ]

                cols_show = list(dict.fromkeys([c for c in cols_show if c and c in detail_calc.columns]))
                detail_affichage = detail_calc[cols_show].copy()

                rename_detail_cols = {}
                if col_client:
                    rename_detail_cols[col_client] = "Client / Référence affaire"
                if col_doc:
                    rename_detail_cols[col_doc] = "N° Document"
                if col_date:
                    rename_detail_cols[col_date] = "Date document"
                if col_agence:
                    rename_detail_cols[col_agence] = "Agence"
                if col_vente:
                    rename_detail_cols[col_vente] = "TOTAL VENTE"
                if col_ca_magasin:
                    rename_detail_cols[col_ca_magasin] = "Vente HT hors acompte"

                detail_affichage = detail_affichage.rename(columns=rename_detail_cols)

                for c in ["TOTAL VENTE", "Vente HT hors acompte", "Remise %", "Bonus / Malus"]:
                    if c in detail_affichage.columns:
                        detail_affichage[c] = pd.to_numeric(detail_affichage[c], errors="coerce").fillna(0).round(2)

                if "Date document" in detail_affichage.columns:
                    detail_affichage["Date document"] = pd.to_datetime(
                        detail_affichage["Date document"],
                        errors="coerce"
                    ).dt.strftime("%d/%m/%Y").fillna("")

                pdf_metrics = {
                    "CA OK": f"{to_float(data.get('ca_ok', 0)):,.2f} EUR",
                    "CA attente": f"{to_float(data.get('ca_attente', 0)):,.2f} EUR",
                    "CA Total": f"{to_float(data.get('ca_total', 0)):,.2f} EUR",
                    "Commission": f"{to_float(data.get('commission_eur', 0)):,.2f} EUR",
                    "Remise commission hors OPC": f"{to_float(data.get('remise_hors_opc_pct', 0)):,.2f} %",
                    "Base commission": f"{to_float(data.get('base_commission_pct', 0)):,.2f} %",
                    "Points perdus": f"{to_float(data.get('points_perdus', 0)):,.0f}",
                    "Commission definitive": f"{to_float(data.get('commission_pct', 0)):,.2f} %",
                    vendeur_remise_label: f"{to_float(data.get(vendeur_remise_col, 0)):,.2f} %",
                    vendeur_bonus_malus_label: f"{to_float(data.get(vendeur_bonus_malus_col, 0)):,.2f} EUR",
                }
                pdf_bytes = make_simple_pdf(
                    f"Detail des affaires - {vendeur} - {periode}",
                    pdf_metrics,
                    detail_affichage.reset_index(drop=True)
                )
                pdf_filename = f"detail_vendeur_{safe_filename(vendeur)}_{safe_filename(periode)}.pdf"
                pdf_container = st.container()
                if is_render_env():
                    pdf_container.download_button(
                        "📄 Télécharger PDF vendeur",
                        pdf_bytes,
                        pdf_filename,
                        "application/pdf",
                        key=f"pdf_vendeur_{safe_filename(vendeur)}",
                        on_click="ignore"
                    )
                else:
                    pdf_col1, pdf_col2 = pdf_container.columns([1, 1])
                    pdf_col1.download_button(
                        "📄 Télécharger PDF vendeur",
                        pdf_bytes,
                        pdf_filename,
                        "application/pdf",
                        key=f"pdf_vendeur_{safe_filename(vendeur)}",
                        on_click="ignore"
                    )
                    if pdf_col2.button("💾 Générer PDF vendeur", key=f"save_pdf_vendeur_{safe_filename(vendeur)}"):
                        pdf_path = save_pdf_export(pdf_bytes, pdf_filename)
                        st.success(f"PDF généré : {pdf_path}")

                def color_bonus_malus(val):
                    try:
                        v = float(val)
                    except Exception:
                        return ""
                    if v > 0:
                        return "background-color: #D1FADF; color: #065F46; font-weight: 700"
                    if v < 0:
                        return "background-color: #FECACA; color: #991B1B; font-weight: 700"
                    return ""

                render_table_with_status_tooltips(
                    detail_affichage,
                    tooltip_by_index=detail_calc["_BULLE_ATTENTE_"].to_dict(),
                    height=600
                )
            else:
                st.info("Aucune affaire trouvée pour ce vendeur.")

    def afficher_agence(tab, agence_forced=None):
        with tab:
            if df_agences.empty:
                st.info("Aucune donnée agence disponible pour ce compte.")
                return

            if agence_forced:
                agence = agence_forced
            else:
                agences_options = sorted(df_agences["agence"])
                agence_precedente = st.session_state.get("agence_select")
                agence_index = agences_options.index(agence_precedente) if agence_precedente in agences_options else 0
                agence = st.selectbox(
                    "Sélectionner une agence",
                    agences_options,
                    index=agence_index,
                    key="agence_select"
                )

            data = df_agences[df_agences["agence"].apply(normalize_key) == normalize_key(agence)]

            if data.empty:
                st.info("Aucune donnée trouvée pour cette agence.")
                return

            data = data.iloc[0]

            c1, c2, c3, c4, c5 = st.columns(5)

            with c1:
                card("✅ CA OK agence", f"{data['ca_ok']:,.2f} €")

            with c2:
                card("⏳ CA en attente agence", f"{data['ca_attente']:,.2f} €")

            with c3:
                card("📌 CA Total agence", f"{data['ca_total']:,.2f} €")

            with c4:
                card("🎯 Bonus / Malus OK", f"{to_float(data.get('bonus_malus_ok', 0)):,.2f} €")

            with c5:
                ratio_opc_agence = to_float(data.get("ratio_opc_total_pct", 0))
                card("🧾 Ratio OPC", f"{ratio_opc_agence:,.1f} %")

            st.write(
                f"Remise moyenne agence : **{data['remise_pct']} %** | "
                f"Affaires OK : **{data['nb_ok']}**"
            )

            st.subheader(f"📋 Détail des affaires de l’agence **{agence}**")

            df_ok = st.session_state.df_ok.copy()
            df_c = st.session_state.df_c.copy()
            col_agence = st.session_state.col_agence
            key_cols = st.session_state.key_cols
            col_client = st.session_state.col_client
            col_doc = st.session_state.col_doc
            col_date = st.session_state.col_date
            col_vente = st.session_state.col_vente
            col_ca_magasin = st.session_state.col_ca_magasin
            col_catalogue = st.session_state.col_catalogue

            ok_detail = df_ok[agence_mask(df_ok, agence, col_agence)].copy()
            attente_detail = df_c[agence_mask(df_c, agence, col_agence)].copy()

            attente_detail = remove_attente_already_ok(
                ok_detail,
                attente_detail,
                key_cols,
                col_client,
                col_agence,
                col_vente,
                col_ca_magasin,
                col_catalogue
            )

            ok_detail["Statut"] = "✅ OK"
            attente_detail["Statut"] = "⏳ En attente"

            detail = pd.concat([ok_detail, attente_detail], ignore_index=True)

            if not detail.empty:
                afficher_alerte_hors_periode(detail, col_client, col_doc, col_date, periode)
                motifs_attente = load_motifs_attente()

                detail["_ATTENTE_KEY_"] = detail.apply(
                    lambda row: make_attente_tracking_key(row, col_client, col_doc, col_agence, col_ca_magasin, col_catalogue),
                    axis=1
                )
                detail["Motif d'attente"] = detail.apply(
                    lambda row: clean_visible(motifs_attente.get(row.get("_ATTENTE_KEY_", ""), {}).get("motif", ""))
                    if str(row.get("Statut", "")).startswith("⏳") else "",
                    axis=1
                )
                detail["Détail motif"] = detail.apply(
                    lambda row: clean_visible(motifs_attente.get(row.get("_ATTENTE_KEY_", ""), {}).get("detail", ""))
                    if str(row.get("Statut", "")).startswith("⏳") else "",
                    axis=1
                )
                detail["_BULLE_ATTENTE_"] = detail.apply(
                    lambda row: " | ".join([
                        part for part in [
                            clean_visible(row.get("Motif d'attente", "")),
                            clean_visible(row.get("Détail motif", ""))
                        ]
                        if part
                    ]),
                    axis=1
                )
                detail["Statut"] = detail.apply(
                    lambda row: "⏳ En attente ⓘ"
                    if clean_visible(row.get("Motif d'attente", "")) else row.get("Statut", ""),
                    axis=1
                )

                colonnes_commerciaux = [
                    st.session_state.col_com1,
                    st.session_state.col_com2,
                    st.session_state.col_com3
                ]
                detail["Commerciaux"] = detail.apply(
                    lambda row: list_commerciaux_row(row, colonnes_commerciaux),
                    axis=1
                )

                cols_show = [
                    st.session_state.col_client,
                    "Commerciaux",
                    st.session_state.col_doc,
                    st.session_state.col_date,
                    "Statut",
                    st.session_state.col_agence,
                    st.session_state.col_ca_magasin,
                    st.session_state.col_vente,
                    st.session_state.col_catalogue,
                    st.session_state.col_rem,
                    st.session_state.col_op
                ]

                cols_show = list(dict.fromkeys([c for c in cols_show if c and c in detail.columns]))
                detail_affichage = detail[cols_show].copy()

                for c in [
                    st.session_state.col_ca_magasin,
                    st.session_state.col_vente,
                    st.session_state.col_catalogue,
                    st.session_state.col_rem
                ]:
                    if c and c in detail_affichage.columns:
                        detail_affichage[c] = pd.to_numeric(detail_affichage[c], errors="coerce").fillna(0).round(2)

                if st.session_state.col_date and st.session_state.col_date in detail_affichage.columns:
                    detail_affichage[st.session_state.col_date] = pd.to_datetime(
                        detail_affichage[st.session_state.col_date],
                        errors="coerce"
                    ).dt.strftime("%d/%m/%Y").fillna("")

                pdf_metrics = {
                    "CA OK agence": f"{to_float(data.get('ca_ok', 0)):,.2f} EUR",
                    "CA attente agence": f"{to_float(data.get('ca_attente', 0)):,.2f} EUR",
                    "CA Total agence": f"{to_float(data.get('ca_total', 0)):,.2f} EUR",
                    "Bonus / Malus OK": f"{to_float(data.get('bonus_malus_ok', 0)):,.2f} EUR",
                    "Remise moyenne": f"{to_float(data.get('remise_pct', 0)):,.2f} %",
                }
                pdf_bytes = make_simple_pdf(
                    f"Detail des affaires agence - {agence} - {periode}",
                    pdf_metrics,
                    detail_affichage.reset_index(drop=True)
                )
                pdf_filename = f"detail_agence_{safe_filename(agence)}_{safe_filename(periode)}.pdf"
                pdf_container = st.container()
                if is_render_env():
                    pdf_container.download_button(
                        "📄 Télécharger PDF agence",
                        pdf_bytes,
                        pdf_filename,
                        "application/pdf",
                        key=f"pdf_agence_{safe_filename(agence)}",
                        on_click="ignore"
                    )
                else:
                    pdf_col1, pdf_col2 = pdf_container.columns([1, 1])
                    pdf_col1.download_button(
                        "📄 Télécharger PDF agence",
                        pdf_bytes,
                        pdf_filename,
                        "application/pdf",
                        key=f"pdf_agence_{safe_filename(agence)}",
                        on_click="ignore"
                    )
                    if pdf_col2.button("💾 Générer PDF agence", key=f"save_pdf_agence_{safe_filename(agence)}"):
                        pdf_path = save_pdf_export(pdf_bytes, pdf_filename)
                        st.success(f"PDF généré : {pdf_path}")

                render_table_with_status_tooltips(
                    detail_affichage,
                    tooltip_by_index=detail["_BULLE_ATTENTE_"].to_dict(),
                    height=600
                )
            else:
                st.info("Aucune affaire trouvée pour cette agence.")

    # ====================== AFFICHAGE SELON ROLE ======================

    if role == "admin":

        if active_page == "📊 Dashboard":
            total_ok = df_vendeurs["ca_ok"].sum()
            total_attente = df_vendeurs["ca_attente"].sum()
            total_global = df_vendeurs["ca_total"].sum()
            total_commissions = df_vendeurs["commission_eur"].sum()
            total_affaires_opc_base = (
                df_vendeurs["nb_total"].sum()
                if "nb_total" in df_vendeurs.columns
                else 0
            )
            total_affaires_opc = (
                df_vendeurs["nb_opc_total"].sum()
                if "nb_opc_total" in df_vendeurs.columns
                else 0
            )
            total_ratio_opc = (
                round(total_affaires_opc / total_affaires_opc_base * 100, 2)
                if total_affaires_opc_base > 0
                else 0.0
            )
            bonus_malus_col = "bonus_malus_ok"
            bonus_malus_label = "🎯 Bonus / Malus OK"
            remise_label = "🎯 Remise OK hors OPC"
            total_bonus_malus = (
                df_vendeurs[bonus_malus_col].sum()
                if bonus_malus_col in df_vendeurs.columns
                else 0
            )
            total_remise = calculate_remise_pct(
                st.session_state.df_ok,
                st.session_state.col_catalogue,
                st.session_state.col_rem,
                st.session_state.col_op
            )

            total_comm_magasin = (
                df_directeurs["commission_magasin_eur"].sum()
                if not df_directeurs.empty and "commission_magasin_eur" in df_directeurs.columns
                else 0
            )

            c1, c2, c3, c4 = st.columns(4)

            with c1:
                card("✅ CA OK vendeurs", f"{total_ok:,.2f} €")

            with c2:
                card("⏳ CA attente vendeurs", f"{total_attente:,.2f} €")

            with c3:
                card("📌 CA Total vendeurs", f"{total_global:,.2f} €")

            with c4:
                card("💰 Commissions vendeurs", f"{total_commissions:,.2f} €")

            c5, c6, c7, c8 = st.columns(4)

            with c5:
                card(remise_label, f"{total_remise:,.2f} %")

            with c6:
                card(bonus_malus_label, f"{total_bonus_malus:,.2f} €")

            with c7:
                card("🏢 Comm. magasin", f"{total_comm_magasin:,.2f} €")

            with c8:
                card("🧾 Ratio OPC", f"{total_ratio_opc:,.1f} %")

            st.divider()

            st.subheader("🏆 Classement vendeurs")
            df_classement_vendeurs = format_df_vendeurs(df_vendeurs)
            df_classement_vendeurs = df_classement_vendeurs.drop(
                columns=[
                    "Bonus / Malus Global",
                    "Remise Global hors OPC",
                    "Remise moy. % hors OPC",
                    "Remise globale % hors OPC"
                ],
                errors="ignore"
            )
            st.dataframe(
                df_classement_vendeurs.sort_values("CA OK", ascending=False).reset_index(drop=True),
                use_container_width=True
            )

            if not df_agences.empty:
                st.subheader("🏢 Classement agences")
                st.dataframe(
                    format_df_agences(df_agences).sort_values("CA OK", ascending=False).reset_index(drop=True),
                    use_container_width=True
                )

        if active_page == "⏳ En attente":
            afficher_dossiers_en_attente(st.container())
        if active_page == "📆 Annuel":
            afficher_annuel(st.container())
        if active_page == "👤 Par Vendeur":
            afficher_vendeur(st.container())
        if active_page == "🏢 Par Agence":
            afficher_agence(st.container())

        if active_page == "👔 Directeurs":
            if df_directeurs.empty:
                st.info("Aucune commission magasin calculée.")
            else:
                st.dataframe(
                    format_df_directeurs(df_directeurs).sort_values("Commission magasin €", ascending=False),
                    use_container_width=True
                )

        if active_page == "📋 Listes complètes":
            st.subheader("👤 Vendeurs")
            st.dataframe(format_df_vendeurs(df_vendeurs).sort_values("CA OK", ascending=False), use_container_width=True)

            st.subheader("🏢 Agences")
            if not df_agences.empty:
                st.dataframe(format_df_agences(df_agences).sort_values("CA OK", ascending=False), use_container_width=True)

            st.subheader("👔 Directeurs")
            if not df_directeurs.empty:
                st.dataframe(format_df_directeurs(df_directeurs).sort_values("Commission magasin €", ascending=False), use_container_width=True)

        if active_page == "⚙️ Utilisateurs":
            afficher_admin_users()

    elif role == "directeur_agence":
        if active_page == "👤 Mes chiffres":
            afficher_vendeur(st.container(), vendeur_forced=user["nom"])
        if active_page == "📆 Annuel":
            afficher_annuel(st.container())
        if active_page == "🏢 Mon agence":
            afficher_agence(st.container(), agence_forced=user["agence"])

        if active_page == "👔 Commission agence":
            if df_directeurs.empty:
                st.info("Aucune commission agence disponible.")
            else:
                st.dataframe(format_df_directeurs(df_directeurs), use_container_width=True)

    elif role == "vendeur":
        if active_page == "👤 Mes chiffres":
            afficher_vendeur(st.container(), vendeur_forced=user["nom"])
        if active_page == "📆 Annuel":
            afficher_annuel(st.container())

else:
    periodes_main = sorted(list_periodes(), key=periode_sort_key)

    if periodes_main:
        st.markdown("""
        <style>
        @keyframes periodChoicePulse {
            0% {
                box-shadow: 0 0 0 0 rgba(54, 135, 23, 0.55);
                border-color: #D1D5DB;
            }
            50% {
                box-shadow: 0 0 0 6px rgba(54, 135, 23, 0.18);
                border-color: #368717;
            }
            100% {
                box-shadow: 0 0 0 0 rgba(54, 135, 23, 0);
                border-color: #D1D5DB;
            }
        }
        .period-choice-panel {
            background: #FFFFFF;
            border: 2px solid #66B32E;
            border-left: 8px solid #66B32E;
            border-radius: 14px;
            padding: 18px 18px 14px 18px;
            box-shadow: 0 4px 14px rgba(102, 179, 46, 0.14);
            margin: 18px 0 20px 0;
        }
        .period-choice-title {
            font-size: 22px;
            font-weight: 800;
            margin-bottom: 4px;
        }
        .period-choice-subtitle {
            color: #4B5563 !important;
            font-size: 14px;
            margin-bottom: 12px;
        }
        .welcome-panel {
            margin: 22px 0 14px 0;
            padding: 18px 22px;
            background: linear-gradient(90deg, #F0F9EB 0%, #FFFFFF 100%);
            border-radius: 14px;
            border: 1px solid #B7E2A0;
        }
        .welcome-title {
            font-size: 30px;
            font-weight: 900;
            letter-spacing: 0;
            margin: 0;
        }
        .welcome-subtitle {
            color: #4B5563 !important;
            font-size: 15px;
            margin-top: 4px;
        }
        div[data-testid="stSelectbox"]:has([data-testid="stWidgetLabel"] label div p) [data-baseweb="select"] {
            animation: periodChoicePulse 2.2s ease-in-out infinite;
            border-radius: 9px;
            border: 2px solid #66B32E !important;
            background: #F7FBF3 !important;
        }
        </style>
        """, unsafe_allow_html=True)
        st.markdown("""
        <div class="welcome-panel">
            <div class="welcome-title">Bienvenue</div>
            <div class="welcome-subtitle">Sélectionne une période pour accéder à tes chiffres EcoHabitat.</div>
        </div>
        """, unsafe_allow_html=True)

        st.markdown("""
        <div class="period-choice-panel">
            <div class="period-choice-title">📅 Choisir une période pour démarrer</div>
            <div class="period-choice-subtitle">Sélectionne une période sauvegardée dans le menu ci-dessous.</div>
        </div>
        """, unsafe_allow_html=True)

        periode_empty_select = st.selectbox(
            "Période sauvegardée",
            [""] + periodes_main,
            key="periode_empty_select",
            help="Choisis une période pour charger le dashboard."
        )

        if periode_empty_select:
            if load_periode_preserve_ui(periode_empty_select):
                st.rerun()
            else:
                st.error("❌ Impossible de charger cette période.")
    else:
        st.warning("Aucune période sauvegardée disponible.")

    if role == "admin":
        st.info("Ou charge les fichiers CONFIRM / BONLIVR puis clique sur « Lancer le traitement ».")

st.caption("✅ Version avec login + gestion utilisateurs • Admin / Vendeur / Directeur agence • Design EcoHabitat • Analyse annuelle M-2")
